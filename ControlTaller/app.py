from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.exceptions import HTTPException
import os
import io
from datetime import date, datetime, timedelta
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from storage import load_workbook_for_app, workbook_to_bytes

app = Flask(__name__,
            template_folder=os.path.join(os.path.dirname(__file__), 'templates'),
            static_folder=os.path.join(os.path.dirname(__file__), 'static'),
            static_url_path='/static')

@app.errorhandler(Exception)
def api_error(error):
    if isinstance(error, HTTPException):
        return error
    app.logger.exception("Error no controlado en ControlTaller")
    return jsonify({"success": False, "error": f"{type(error).__name__}: {error}"}), 500

EXCEL_FILE = "google-sheets://ControlTaller"

COMISION_DEFAULT = 50  # % de la mano de obra que se paga al mecánico (editable)

CATEGORIAS_GASTO_DEFAULT = ["Gasolina", "Repuestos", "Almuerzos", "Herramientas", "Servicios", "Otros"]

SHEET_HEADERS = {
    "Mecanicos": ["ID", "Nombre", "Teléfono", "% Comisión", "Activo", "Fecha Ingreso"],
    "Equipos": ["ID", "Nombre", "Integrante 1", "Integrante 2", "% Comisión Total", "Activo"],
    "Trabajos": ["ID", "Fecha", "Semana", "Mecánico", "Equipo", "Placa", "Vehículo",
                 "Descripción", "Monto Mano de Obra", "% Comisión", "Monto Mecánico",
                 "Factura Cancelada", "Estado", "Fecha Pago", "Grupo"],
    "Gastos": ["ID", "Fecha", "Categoría", "Descripción", "Monto", "Responsable", "Método de Pago"],
    "Pagos": ["ID", "Fecha Pago", "Mecánico", "Semana", "N° Trabajos",
              "Total Mano de Obra", "Total Comisión", "Total Descuentos", "Neto Pagado"],
    "Prestamos": ["ID", "Fecha", "Mecánico", "Monto Original", "Cuota Sugerida",
                  "Total Descontado", "Saldo Pendiente", "Estado", "Observaciones"],
    "Descuentos Nomina": ["ID", "Fecha Aplicación", "Mecánico", "Semana", "Concepto",
                          "Monto", "Préstamo ID", "Observaciones"],
    "Deudas Taller": ["ID", "Fecha Registro", "Acreedor", "Concepto", "Monto Total", "Frecuencia",
                      "Día Pago", "Próximo Vencimiento", "Estado", "Observaciones", "Tipo"],
    "Fondos Deudas": ["ID", "Fecha Aporte", "Deuda ID", "Período", "Acreedor", "Monto", "Método", "Observaciones"],
    "Pagos Deudas": ["ID", "Fecha Pago", "Deuda ID", "Período", "Acreedor", "Monto", "Tipo Pago", "Observaciones"],
    "Herramientas": ["ID", "Herramienta", "Prestada A", "Entregada Por", "Fecha Préstamo",
                      "Fecha Devolución", "Estado", "Observaciones"],
}

# ---------------------------------------------------------------------------
# Las pestañas y encabezados remotos se preparan en storage.py.
# Este módulo solo trabaja sobre la instantánea recibida de Google Sheets.

SHEETS_NOMINA = ("Trabajos", "Descuentos Nomina")
SHEETS_PRESTAMOS = ("Prestamos", "Descuentos Nomina")
SHEETS_PRESTAMOS_PANEL = ("Trabajos", "Prestamos", "Descuentos Nomina")
SHEETS_PRESTAMO_WRITE = ("Mecanicos", "Prestamos")
SHEETS_DEUDAS = ("Deudas Taller", "Fondos Deudas", "Pagos Deudas")
SHEETS_PAGAR_NOMINA = ("Trabajos", "Descuentos Nomina", "Pagos")


def get_wb(sheet_titles=None, force_refresh=False):
    """Carga una instantánea actual, limitada a las pestañas requeridas."""
    return load_workbook_for_app(
        EXCEL_FILE,
        sheet_titles=sheet_titles,
        force_refresh=force_refresh,
    )


def next_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (TypeError, ValueError):
                pass
    return max_id + 1


def append_row(ws, headers, row_values, id_val, zebra=True):
    row_num = ws.max_row + 1
    fill_color = "F9EBEA" if zebra and row_num % 2 == 0 else "FFFFFF"
    for col, header in enumerate(headers, 1):
        value = row_values.get(header, "")
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        if col == 1:
            cell.font = Font(bold=True, color="C0392B")
    return row_num


def sheet_to_dicts(ws, headers):
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            result.append(dict(zip(headers, row)))
    return result


def find_row_by_id(ws, target_id):
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and str(row[0].value) == str(target_id):
            return row
    return None


def week_key(fecha_str):
    """fecha_str = 'YYYY-MM-DD' -> 'YYYY-Wnn' (semana ISO)"""
    d = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def money(value):
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def normalize_name(value):
    return " ".join(str(value or "").strip().split()).casefold()


def _row_values(row, headers):
    return {headers[i]: row[i].value for i in range(min(len(headers), len(row)))}


def _loan_records(wb):
    """Devuelve préstamos con saldo calculado y no altera la hoja."""
    ws = wb["Prestamos"]
    headers = SHEET_HEADERS["Prestamos"]
    records = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        values = _row_values(row, headers)
        original = money(values.get("Monto Original"))
        discounted = money(values.get("Total Descontado"))
        saldo = max(round(original - discounted, 2), 0.0)
        estado = "Pagado" if saldo <= 0.009 else "Pendiente"
        records.append({
            "id": int(values.get("ID") or 0),
            "fecha": values.get("Fecha") or "",
            "mecanico": values.get("Mecánico") or "",
            "monto_original": original,
            "cuota_sugerida": money(values.get("Cuota Sugerida")),
            "total_descontado": discounted,
            "saldo_pendiente": saldo,
            "estado": estado,
            "observaciones": values.get("Observaciones") or "",
        })
    return records


def _discount_records(wb):
    ws = wb["Descuentos Nomina"]
    headers = SHEET_HEADERS["Descuentos Nomina"]
    records = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        values = _row_values(row, headers)
        records.append({
            "id": int(values.get("ID") or 0),
            "fecha_aplicacion": values.get("Fecha Aplicación") or "",
            "mecanico": values.get("Mecánico") or "",
            "semana": values.get("Semana") or "",
            "concepto": values.get("Concepto") or "",
            "monto": money(values.get("Monto")),
            "prestamo_id": int(values["Préstamo ID"]) if str(values.get("Préstamo ID") or "").strip().isdigit() else None,
            "observaciones": values.get("Observaciones") or "",
        })
    return records


def _nomina_calculada(wb, semana=None, mecanico=None):
    """Calcula bruto pendiente, descuentos aplicados y neto por mecánico."""
    ws = wb["Trabajos"]
    headers = SHEET_HEADERS["Trabajos"]
    resumen = {}
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        values = _row_values(row, headers)
        if semana and values.get("Semana") != semana:
            continue
        if mecanico and values.get("Mecánico") != mecanico:
            continue
        if values.get("Factura Cancelada") != "Sí":
            continue
        mec = values.get("Mecánico") or "Sin asignar"
        item = resumen.setdefault(mec, {
            "mecanico": mec, "n_trabajos": 0, "total_mo": 0.0,
            "total_comision": 0.0, "pendiente": 0.0, "pagado": 0.0,
            "bruto_pendiente": 0.0, "descuentos": 0.0, "neto_pagar": 0.0,
        })
        monto_mo = money(values.get("Monto Mano de Obra"))
        monto_mec = money(values.get("Monto Mecánico"))
        item["n_trabajos"] += 1
        item["total_mo"] += monto_mo
        item["total_comision"] += monto_mec
        if values.get("Estado") == "Pagado":
            item["pagado"] += monto_mec
        else:
            item["pendiente"] += monto_mec

    discounts = _discount_records(wb)
    for item in resumen.values():
        if semana:
            item["descuentos"] = sum(
                d["monto"] for d in discounts
                if d["mecanico"] == item["mecanico"] and d["semana"] == semana
            )
        else:
            item["descuentos"] = sum(
                d["monto"] for d in discounts if d["mecanico"] == item["mecanico"]
            )
        item["bruto_pendiente"] = round(item["pendiente"], 2)
        item["descuentos"] = round(item["descuentos"], 2)
        item["neto_pagar"] = round(max(item["bruto_pendiente"] - item["descuentos"], 0), 2)
        for key in ("total_mo", "total_comision", "pendiente", "pagado"):
            item[key] = round(item[key], 2)
    return list(resumen.values())


def _nomina_response(wb, semana=None, mecanico=None):
    resumen = _nomina_calculada(wb, semana=semana, mecanico=mecanico)
    ws = wb["Trabajos"]
    headers = SHEET_HEADERS["Trabajos"]
    semanas = sorted({
        _row_values(row, headers).get("Semana")
        for row in ws.iter_rows(min_row=2)
        if row[0].value is not None and _row_values(row, headers).get("Semana")
    }, reverse=True)
    totales = {
        "bruto_pendiente": round(sum(r["bruto_pendiente"] for r in resumen), 2),
        "descuentos": round(sum(r["descuentos"] for r in resumen), 2),
        "neto_pagar": round(sum(r["neto_pagar"] for r in resumen), 2),
    }
    return {"resumen": resumen, "semanas_disponibles": semanas, "totales": totales}


# ---------------------------------------------------------------------------
# Rutas de página
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template('index.html', comision_default=COMISION_DEFAULT,
                            categorias=CATEGORIAS_GASTO_DEFAULT)


# ---------------------------------------------------------------------------
# API Mecánicos
# ---------------------------------------------------------------------------

@app.route('/api/mecanicos', methods=['GET'])
def list_mecanicos():
    wb = get_wb()
    ws = wb["Mecanicos"]
    return jsonify(sheet_to_dicts(ws, SHEET_HEADERS["Mecanicos"]))


@app.route('/api/mecanicos', methods=['POST'])
def create_mecanico():
    data = request.json
    wb = get_wb()
    ws = wb["Mecanicos"]
    mid = next_id(ws)
    append_row(ws, SHEET_HEADERS["Mecanicos"], {
        "ID": mid,
        "Nombre": data.get("nombre", "").strip(),
        "Teléfono": data.get("telefono", ""),
        "% Comisión": float(data.get("comision", COMISION_DEFAULT)),
        "Activo": "Sí",
        "Fecha Ingreso": data.get("fecha_ingreso", datetime.now().strftime("%Y-%m-%d")),
    }, mid)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True, "id": mid})


@app.route('/api/mecanicos/<int:mid>', methods=['PUT'])
def update_mecanico(mid):
    data = request.json
    wb = get_wb()
    ws = wb["Mecanicos"]
    row = find_row_by_id(ws, mid)
    if not row:
        return jsonify({"success": False, "error": "No encontrado"}), 404
    headers = SHEET_HEADERS["Mecanicos"]
    field_map = {"nombre": "Nombre", "telefono": "Teléfono",
                 "comision": "% Comisión", "activo": "Activo"}
    for key, header in field_map.items():
        if key in data:
            col = headers.index(header) + 1
            val = data[key]
            if key == "comision":
                val = float(val)
            ws.cell(row=row[0].row, column=col, value=val)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/mecanicos/<int:mid>', methods=['DELETE'])
def delete_mecanico(mid):
    wb = get_wb()
    ws = wb["Mecanicos"]
    row = find_row_by_id(ws, mid)
    if row:
        ws.delete_rows(row[0].row, 1)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# API Equipos (mecánicos en pareja)
# ---------------------------------------------------------------------------

@app.route('/api/equipos', methods=['GET'])
def list_equipos():
    wb = get_wb()
    ws = wb["Equipos"]
    return jsonify(sheet_to_dicts(ws, SHEET_HEADERS["Equipos"]))


@app.route('/api/equipos', methods=['POST'])
def create_equipo():
    data = request.json
    m1 = (data.get("integrante1") or "").strip()
    m2 = (data.get("integrante2") or "").strip()
    if not m1 or not m2 or m1 == m2:
        return jsonify({"success": False, "error": "Selecciona dos mecánicos distintos"}), 400
    wb = get_wb()
    ws = wb["Equipos"]
    eid = next_id(ws)
    append_row(ws, SHEET_HEADERS["Equipos"], {
        "ID": eid,
        "Nombre": f"{m1} + {m2}",
        "Integrante 1": m1,
        "Integrante 2": m2,
        "% Comisión Total": float(data.get("comision_total", COMISION_DEFAULT)),
        "Activo": "Sí",
    }, eid)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True, "id": eid})


@app.route('/api/equipos/<int:eid>', methods=['PUT'])
def update_equipo(eid):
    data = request.json
    wb = get_wb()
    ws = wb["Equipos"]
    row = find_row_by_id(ws, eid)
    if not row:
        return jsonify({"success": False, "error": "No encontrado"}), 404
    headers = SHEET_HEADERS["Equipos"]
    field_map = {"comision_total": "% Comisión Total", "activo": "Activo"}
    for key, header in field_map.items():
        if key in data:
            col = headers.index(header) + 1
            val = data[key]
            if key == "comision_total":
                val = float(val)
            ws.cell(row=row[0].row, column=col, value=val)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/equipos/<int:eid>', methods=['DELETE'])
def delete_equipo(eid):
    wb = get_wb()
    ws = wb["Equipos"]
    row = find_row_by_id(ws, eid)
    if row:
        ws.delete_rows(row[0].row, 1)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# API Trabajos / Nómina
# ---------------------------------------------------------------------------

@app.route('/api/trabajos', methods=['GET'])
def list_trabajos():
    wb = get_wb()
    ws = wb["Trabajos"]
    data = sheet_to_dicts(ws, SHEET_HEADERS["Trabajos"])
    semana = request.args.get('semana')
    mecanico = request.args.get('mecanico')
    if semana:
        data = [d for d in data if d.get("Semana") == semana]
    if mecanico:
        data = [d for d in data if d.get("Mecánico") == mecanico]
    data.sort(key=lambda d: (d.get("Fecha") or "", d.get("ID") or 0), reverse=True)
    return jsonify(data)


@app.route('/api/trabajo', methods=['POST'])
def create_trabajo():
    data = request.json
    fecha = data.get("fecha", datetime.now().strftime("%Y-%m-%d"))
    monto_mo = float(data.get("monto_mo", 0) or 0)
    semana = week_key(fecha)
    placa = (data.get("placa", "") or "").upper()
    vehiculo = data.get("vehiculo", "")
    descripcion = data.get("descripcion", "")
    factura_cancelada = "Sí" if data.get("factura_cancelada") else "No"

    print(f"\n[DEBUG] POST /api/trabajo recibido: {data}")
    print(f"[DEBUG] data.get('factura_cancelada') = {data.get('factura_cancelada')!r} (tipo {type(data.get('factura_cancelada')).__name__}) -> se guardará como: {factura_cancelada!r}\n")

    wb = get_wb()
    ws = wb["Trabajos"]

    base = {
        "Fecha": fecha, "Semana": semana, "Placa": placa, "Vehículo": vehiculo,
        "Descripción": descripcion, "Monto Mano de Obra": monto_mo,
        "Factura Cancelada": factura_cancelada,
        "Estado": "Pendiente", "Fecha Pago": "",
    }

    if data.get("modo") == "equipo":
        # --- Trabajo en pareja: se reparte el % total entre los dos integrantes ---
        equipo_id = data.get("equipo_id")
        ws_eq = wb["Equipos"]
        row_eq = find_row_by_id(ws_eq, equipo_id)
        if not row_eq:
            return jsonify({"success": False, "error": "Equipo no encontrado"}), 404
        eq_headers = SHEET_HEADERS["Equipos"]
        eq_vals = {eq_headers[i]: row_eq[i].value for i in range(len(eq_headers))}

        comision_total = float(data.get("comision_pct", eq_vals.get("% Comisión Total", COMISION_DEFAULT)) or COMISION_DEFAULT)
        comision_individual = round(comision_total / 2, 4)
        monto_individual = round(monto_mo * comision_individual / 100, 2)
        nombre_equipo = eq_vals.get("Nombre", "")
        integrantes = [eq_vals.get("Integrante 1"), eq_vals.get("Integrante 2")]

        grupo_id = next_id(ws)
        for i, mec in enumerate(integrantes):
            tid = grupo_id if i == 0 else next_id(ws)
            row_data = dict(base)
            row_data.update({
                "ID": tid, "Mecánico": mec, "Equipo": nombre_equipo,
                "% Comisión": comision_individual, "Monto Mecánico": monto_individual,
                "Grupo": grupo_id,
            })
            append_row(ws, SHEET_HEADERS["Trabajos"], row_data, tid)

        wb.save(EXCEL_FILE)
        return jsonify({"success": True, "monto_mecanico": monto_individual, "equipo": nombre_equipo,
                         "comision_individual": comision_individual})

    else:
        # --- Trabajo individual ---
        mecanico = data.get("mecanico", "")
        comision_pct = float(data.get("comision_pct", COMISION_DEFAULT) or COMISION_DEFAULT)
        monto_mecanico = round(monto_mo * comision_pct / 100, 2)
        tid = next_id(ws)
        row_data = dict(base)
        row_data.update({
            "ID": tid, "Mecánico": mecanico, "Equipo": "",
            "% Comisión": comision_pct, "Monto Mecánico": monto_mecanico, "Grupo": "",
        })
        append_row(ws, SHEET_HEADERS["Trabajos"], row_data, tid)
        wb.save(EXCEL_FILE)
        return jsonify({"success": True, "id": tid, "monto_mecanico": monto_mecanico})


@app.route('/api/trabajo/<int:tid>', methods=['PUT'])
def update_trabajo(tid):
    data = request.json
    wb = get_wb()
    ws = wb["Trabajos"]
    row = find_row_by_id(ws, tid)
    if not row:
        return jsonify({"success": False, "error": "No encontrado"}), 404
    headers = SHEET_HEADERS["Trabajos"]
    r = row[0].row

    def set_val(header, value):
        ws.cell(row=r, column=headers.index(header) + 1, value=value)

    if "monto_mo" in data or "comision_pct" in data:
        monto_mo = float(data.get("monto_mo", ws.cell(row=r, column=headers.index("Monto Mano de Obra")+1).value or 0))
        comision_pct = float(data.get("comision_pct", ws.cell(row=r, column=headers.index("% Comisión")+1).value or COMISION_DEFAULT))
        set_val("Monto Mano de Obra", monto_mo)
        set_val("% Comisión", comision_pct)
        set_val("Monto Mecánico", round(monto_mo * comision_pct / 100, 2))
    for key, header in {"descripcion": "Descripción", "placa": "Placa",
                         "vehiculo": "Vehículo"}.items():
        if key in data:
            set_val(header, data[key])
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/trabajo/<int:tid>/factura', methods=['POST'])
def marcar_factura_cancelada(tid):
    wb = get_wb()
    ws = wb["Trabajos"]
    headers = SHEET_HEADERS["Trabajos"]
    row = find_row_by_id(ws, tid)
    if not row:
        return jsonify({"success": False, "error": "No encontrado"}), 404

    # Si el trabajo es de un equipo (Grupo), se marca para ambos integrantes a la vez.
    grupo = row[headers.index("Grupo")].value
    filas_a_marcar = [row[0].row]
    if grupo:
        for r in ws.iter_rows(min_row=2):
            if r[0].value is None or r[0].row == row[0].row:
                continue
            if r[headers.index("Grupo")].value == grupo:
                filas_a_marcar.append(r[0].row)

    col = headers.index("Factura Cancelada") + 1
    for r_num in filas_a_marcar:
        ws.cell(row=r_num, column=col, value="Sí")

    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/trabajo/<int:tid>', methods=['DELETE'])
def delete_trabajo(tid):
    wb = get_wb()
    ws = wb["Trabajos"]
    headers = SHEET_HEADERS["Trabajos"]
    row = find_row_by_id(ws, tid)
    if not row:
        wb.save(EXCEL_FILE)
        return jsonify({"success": True})

    grupo = row[headers.index("Grupo")].value
    rows_to_delete = [row[0].row]

    if grupo:
        for r in ws.iter_rows(min_row=2):
            if r[0].value is None or r[0].row == row[0].row:
                continue
            if r[headers.index("Grupo")].value == grupo:
                rows_to_delete.append(r[0].row)

    for r_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r_num, 1)

    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/nomina/resumen', methods=['GET'])
def nomina_resumen():
    semana = request.args.get('semana')
    wb = get_wb(SHEETS_NOMINA)
    return jsonify(_nomina_response(wb, semana=semana))


# ---------------------------------------------------------------------------
# API Préstamos y descuentos de Nómina
# ---------------------------------------------------------------------------

@app.route('/api/prestamos', methods=['GET'])
def list_prestamos():
    wb = get_wb(SHEETS_PRESTAMOS)
    mecanico = request.args.get('mecanico')
    prestamos = _loan_records(wb)
    descuentos = _discount_records(wb)
    if mecanico:
        prestamos = [p for p in prestamos if p['mecanico'] == mecanico]
        descuentos = [d for d in descuentos if d['mecanico'] == mecanico]
    return jsonify({'prestamos': prestamos, 'descuentos': descuentos})


@app.route('/api/prestamos/panel', methods=['GET'])
def prestamos_panel():
    semana = request.args.get('semana')
    wb = get_wb(SHEETS_PRESTAMOS_PANEL)
    return jsonify({
        'nomina': _nomina_response(wb, semana=semana),
        'prestamos': _loan_records(wb),
        'descuentos': _discount_records(wb),
    })


@app.route('/api/prestamo', methods=['POST'])
def create_prestamo():
    data = request.json or {}
    mecanico = ' '.join(str(data.get('mecanico') or '').strip().split())
    try:
        monto = round(float(data.get('monto_original') or 0), 2)
        cuota = round(float(data.get('cuota_sugerida') or 0), 2)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'El monto y la cuota deben ser numéricos'}), 400
    if not mecanico or monto <= 0 or cuota <= 0:
        return jsonify({'success': False, 'error': 'Mecánico, monto y cuota son obligatorios y deben ser mayores que cero'}), 400
    if cuota > monto:
        cuota = monto

    wb = get_wb(SHEETS_PRESTAMO_WRITE)
    ws_mec = wb['Mecanicos']
    if not any(normalize_name(row[1].value) == normalize_name(mecanico) and row[4].value != 'No'
               for row in ws_mec.iter_rows(min_row=2) if row[0].value is not None):
        return jsonify({'success': False, 'error': 'El mecánico no existe o está inactivo'}), 400
    ws = wb['Prestamos']
    pid = next_id(ws)
    fecha = data.get('fecha') or datetime.now().strftime('%Y-%m-%d')
    append_row(ws, SHEET_HEADERS['Prestamos'], {
        'ID': pid, 'Fecha': fecha, 'Mecánico': mecanico,
        'Monto Original': monto, 'Cuota Sugerida': cuota,
        'Total Descontado': 0.0, 'Saldo Pendiente': monto,
        'Estado': 'Pendiente', 'Observaciones': (data.get('observaciones') or '').strip(),
    }, pid)
    wb.save(EXCEL_FILE)
    return jsonify({'success': True, 'id': pid, 'saldo_pendiente': monto})


@app.route('/api/prestamo/<int:pid>', methods=['PUT'])
def update_prestamo(pid):
    data = request.json or {}
    wb = get_wb(SHEETS_PRESTAMOS)
    ws = wb['Prestamos']
    row = find_row_by_id(ws, pid)
    if not row:
        return jsonify({'success': False, 'error': 'Préstamo no encontrado'}), 404
    headers = SHEET_HEADERS['Prestamos']
    original = money(row[headers.index('Monto Original')].value)
    descontado = money(row[headers.index('Total Descontado')].value)
    if 'cuota_sugerida' in data:
        try:
            cuota = round(float(data.get('cuota_sugerida') or 0), 2)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'La cuota debe ser numérica'}), 400
        if cuota <= 0 or cuota > max(round(original - descontado, 2), 0):
            return jsonify({'success': False, 'error': 'La cuota debe ser mayor que cero y no superar el saldo pendiente'}), 400
        ws.cell(row=row[0].row, column=headers.index('Cuota Sugerida') + 1, value=cuota)
    if 'observaciones' in data:
        ws.cell(row=row[0].row, column=headers.index('Observaciones') + 1, value=(data.get('observaciones') or '').strip())
    wb.save(EXCEL_FILE)
    return jsonify({'success': True})


@app.route('/api/prestamo/<int:pid>', methods=['DELETE'])
def delete_prestamo(pid):
    wb = get_wb(SHEETS_PRESTAMOS)
    ws = wb['Prestamos']
    row = find_row_by_id(ws, pid)
    if not row:
        return jsonify({'success': False, 'error': 'Préstamo no encontrado'}), 404
    if any(d['prestamo_id'] == pid for d in _discount_records(wb)):
        return jsonify({'success': False, 'error': 'No se puede eliminar un préstamo que ya tiene descuentos aplicados'}), 409
    ws.delete_rows(row[0].row, 1)
    wb.save(EXCEL_FILE)
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# API Deudas del Taller, fondos y pagos
# ---------------------------------------------------------------------------

def _parse_iso_date(value, field='fecha'):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or '').strip()
    if not text:
        raise ValueError(f'{field} es obligatoria')
    try:
        return datetime.strptime(text[:10], '%Y-%m-%d').date()
    except ValueError as exc:
        raise ValueError(f'{field} debe tener formato AAAA-MM-DD') from exc


def _frequency(value):
    normalized = normalize_name(value)
    if normalized in {'mensual', 'monthly'}:
        return 'Mensual'
    if normalized in {'semanal', 'weekly'}:
        return 'Semanal'
    if normalized in {'', 'unico', 'único', 'unica', 'única', 'pago unico', 'pago único', 'one time'}:
        return 'Único'
    raise ValueError('La frecuencia debe ser Único, Semanal o Mensual')


def _debt_type(value, frequency=None):
    """Normaliza el tipo nuevo e infiere el tipo de filas antiguas."""
    normalized = normalize_name(value)
    if normalized in {'recurrente', 'recurrent', 'fijo', 'fija'}:
        tipo = 'Recurrente'
    elif normalized in {'variable'}:
        tipo = 'Variable'
    elif not normalized:
        # Compatibilidad: Mensual/Semanal eran recurrentes; Único era una
        # obligación que se cerraba al pagarla completamente.
        tipo = 'Recurrente' if frequency in {'Mensual', 'Semanal'} else 'Variable'
    else:
        raise ValueError('El tipo debe ser Recurrente o Variable')
    if tipo == 'Recurrente' and frequency == 'Único':
        raise ValueError('Una deuda recurrente debe tener frecuencia Mensual o Semanal')
    return tipo


def _safe_month_day(year, month, day_number):
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(int(day_number), last_day))


def _next_month_day(base_date, day_number):
    if base_date.month == 12:
        year, month = base_date.year + 1, 1
    else:
        year, month = base_date.year, base_date.month + 1
    return _safe_month_day(year, month, day_number)


def _monthly_due(day_number, reference=None):
    reference = reference or date.today()
    candidate = _safe_month_day(reference.year, reference.month, day_number)
    if candidate < reference:
        candidate = _next_month_day(reference, day_number)
    return candidate


def _next_week_day(reference):
    return reference + timedelta(days=7)


def _period_start(frecuencia, vencimiento):
    """Devuelve el inicio del período que corresponde al próximo vencimiento."""
    if frecuencia == 'Mensual':
        return date(vencimiento.year, vencimiento.month, 1)
    if frecuencia == 'Semanal':
        return vencimiento - timedelta(days=6)
    return vencimiento


def _fondos_deuda(wb):
    ws = wb['Fondos Deudas']
    headers = SHEET_HEADERS['Fondos Deudas']
    records = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        values = _row_values(row, headers)
        raw_debt = values.get('Deuda ID')
        records.append({
            'id': int(values.get('ID') or 0),
            'fecha': values.get('Fecha Aporte') or '',
            'deuda_id': int(raw_debt) if str(raw_debt or '').strip().isdigit() else None,
            'periodo': str(values.get('Período') or '').strip(),
            'acreedor': values.get('Acreedor') or '',
            'monto': money(values.get('Monto')),
            'metodo': values.get('Método') or '',
            'observaciones': values.get('Observaciones') or '',
        })
    return records


def _pagos_deuda(wb):
    ws = wb['Pagos Deudas']
    headers = SHEET_HEADERS['Pagos Deudas']
    records = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        values = _row_values(row, headers)
        raw_debt = values.get('Deuda ID')
        records.append({
            'id': int(values.get('ID') or 0),
            'fecha': values.get('Fecha Pago') or '',
            'deuda_id': int(raw_debt) if str(raw_debt or '').strip().isdigit() else None,
            'periodo': str(values.get('Período') or '').strip(),
            'acreedor': values.get('Acreedor') or '',
            'monto': money(values.get('Monto')),
            'tipo_pago': values.get('Tipo Pago') or '',
            'observaciones': values.get('Observaciones') or '',
        })
    return records


def _deuda_records(wb):
    ws = wb['Deudas Taller']
    headers = SHEET_HEADERS['Deudas Taller']
    fondos = _fondos_deuda(wb)
    pagos = _pagos_deuda(wb)
    hoy = date.today()
    records = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        values = _row_values(row, headers)
        deuda_id = int(values.get('ID') or 0)
        try:
            frecuencia_base = _frequency(values.get('Frecuencia'))
        except ValueError:
            frecuencia_base = 'Único'
        try:
            tipo = _debt_type(values.get('Tipo'), frecuencia_base)
        except ValueError:
            tipo = 'Variable'
        frecuencia = frecuencia_base if tipo == 'Recurrente' else 'Único'
        monto_total = money(values.get('Monto Total'))
        try:
            vencimiento = _parse_iso_date(values.get('Próximo Vencimiento'), 'Próximo vencimiento')
        except ValueError:
            vencimiento = hoy
        if frecuencia == 'Mensual':
            periodo = vencimiento.strftime('%Y-%m')
        elif frecuencia == 'Semanal':
            periodo = vencimiento.isoformat()
        else:
            periodo = 'Único'
        periodo_inicio = _period_start(frecuencia, vencimiento)
        periodo_iniciado = frecuencia == 'Único' or hoy >= periodo_inicio
        fondos_deuda = [f for f in fondos if f['deuda_id'] == deuda_id]
        pagos_deuda = [p for p in pagos if p['deuda_id'] == deuda_id]
        fondos_periodo = sum(f['monto'] for f in fondos_deuda if frecuencia == 'Único' or f['periodo'] == periodo)
        pagos_periodo = sum(p['monto'] for p in pagos_deuda if frecuencia == 'Único' or p['periodo'] == periodo)
        aportado_total = round(sum(f['monto'] for f in fondos_deuda), 2)
        pagado_total = round(sum(p['monto'] for p in pagos_deuda), 2)
        saldo_periodo = round(max(monto_total - pagos_periodo, 0.0), 2)
        fondo_disponible = round(max(fondos_periodo - pagos_periodo, 0.0), 2)
        faltante = round(max(saldo_periodo - fondo_disponible, 0.0), 2)
        estado_base = str(values.get('Estado') or '').strip()
        if not periodo_iniciado and frecuencia in {'Mensual', 'Semanal'}:
            # El período siguiente queda programado, pero no es una obligación
            # vigente hasta que comienza su ciclo. No debe sumarse al total actual.
            saldo_periodo = 0.0
            fondo_disponible = 0.0
            faltante = 0.0
            estado = 'Programada'
        elif frecuencia == 'Único' and saldo_periodo <= 0.009:
            estado = 'Pagado'
        elif frecuencia in {'Mensual', 'Semanal'} and saldo_periodo <= 0.009:
            estado = 'Pagado'
        elif vencimiento < hoy:
            estado = 'Atrasado'
        elif fondo_disponible >= saldo_periodo and saldo_periodo > 0:
            estado = 'Listo para pagar'
        else:
            estado = 'Pendiente de reunir'
        if estado_base == 'Pagado' and frecuencia == 'Único' and saldo_periodo <= 0.009:
            estado = 'Pagado'
        records.append({
            'id': deuda_id,
            'fecha_registro': values.get('Fecha Registro') or '',
            'acreedor': values.get('Acreedor') or '',
            'concepto': values.get('Concepto') or '',
            'monto_total': monto_total,
            'frecuencia': frecuencia,
            'tipo': tipo,
            'dia_pago': int(values.get('Día Pago') or vencimiento.day),
            'proximo_vencimiento': vencimiento.isoformat(),
            'periodo': periodo,
            'periodo_inicio': periodo_inicio.isoformat(),
            'periodo_iniciado': periodo_iniciado,
            'estado': estado,
            'observaciones': values.get('Observaciones') or '',
            'fondos_aportados': aportado_total,
            'fondo_disponible': fondo_disponible,
            'pagado_total': pagado_total,
            'saldo_pendiente': saldo_periodo,
            'faltante_reunir': faltante,
            'fondos_periodo': round(fondos_periodo, 2),
            'pagado_periodo': round(pagos_periodo, 2),
        })
    return records


def _payment_month(fecha):
    """Devuelve YYYY-MM para una fecha de pago remota o local."""
    if isinstance(fecha, (date, datetime)):
        return fecha.strftime('%Y-%m')
    text = str(fecha or '').strip()
    return text[:7] if len(text) >= 7 and text[4] == '-' and text[5:7].isdigit() else ''


def _deudas_panel(wb):
    deudas = _deuda_records(wb)
    fondos = _fondos_deuda(wb)
    pagos = _pagos_deuda(wb)
    hoy = date.today()
    mes_actual = hoy.strftime('%Y-%m')
    # Las Variables pagadas dejan de ser compromisos activos, pero siguen
    # en `deudas` y en `pagos` para conservar el registro histórico.
    deudas_activas = [
        d for d in deudas
        if not (d['tipo'] == 'Variable' and d['saldo_pendiente'] <= 0.009)
        and not (d['tipo'] == 'Recurrente' and not d['periodo_iniciado'])
    ]
    deudas_programadas = [
        d for d in deudas
        if d['tipo'] == 'Recurrente' and not d['periodo_iniciado']
    ]
    calendario = []
    for deuda in deudas_activas:
        vencimiento = _parse_iso_date(deuda['proximo_vencimiento'], 'Próximo vencimiento')
        calendario.append({
            'deuda_id': deuda['id'],
            'acreedor': deuda['acreedor'],
            'concepto': deuda['concepto'],
            'monto': deuda['saldo_pendiente'],
            'fondos': deuda['fondo_disponible'],
            'faltante': deuda['faltante_reunir'],
            'fecha': vencimiento.isoformat(),
            'dias': (vencimiento - hoy).days,
            'estado': deuda['estado'],
        })
    calendario.sort(key=lambda item: (item['fecha'], item['acreedor']))

    balance = {}
    for pago in pagos:
        mes = _payment_month(pago.get('fecha'))
        if not mes:
            continue
        item = balance.setdefault(mes, {'mes': mes, 'pagado': 0.0, 'operaciones': 0})
        item['pagado'] = round(item['pagado'] + pago['monto'], 2)
        item['operaciones'] += 1
    balance_mensual = sorted(balance.values(), key=lambda item: item['mes'], reverse=True)
    pagado_total = round(sum(p['monto'] for p in pagos), 2)
    pagado_mes = round(sum(p['monto'] for p in pagos if _payment_month(p.get('fecha')) == mes_actual), 2)
    # Solo obligaciones activas con saldo materialmente pendiente; las Variables
    # pagadas permanecen en el histórico, pero no deben inflar este indicador.
    falta_recurrente = round(sum(
        d['saldo_pendiente'] for d in deudas_activas
        if d['tipo'] == 'Recurrente' and d['saldo_pendiente'] > 0.009
    ), 2)
    falta_variable = round(sum(
        d['saldo_pendiente'] for d in deudas_activas
        if d['tipo'] == 'Variable' and d['saldo_pendiente'] > 0.009
    ), 2)
    falta_pagar = round(falta_recurrente + falta_variable, 2)

    return {
        'deudas': deudas,
        'deudas_activas': deudas_activas,
        'deudas_programadas': deudas_programadas,
        'fondos': fondos,
        'pagos': pagos,
        'balance_mensual': balance_mensual,
        'calendario': calendario,
        'hoy': hoy.isoformat(),
        'totales': {
            'comprometido': falta_pagar,
            'faltante_pagar': falta_pagar,
            'faltante_recurrente': falta_recurrente,
            'faltante_variable': falta_variable,
            'pagado_total': pagado_total,
            'pagado_mes': pagado_mes,
            'mes_actual': mes_actual,
            'fondos_disponibles': round(sum(d['fondo_disponible'] for d in deudas_activas), 2),
            'faltante_reunir': round(sum(d['faltante_reunir'] for d in deudas_activas), 2),
            'proximos': sum(1 for item in calendario if 0 <= item['dias'] <= 7),
            'atrasados': sum(1 for item in calendario if item['dias'] < 0),
        },
    }


def _get_deuda_summary(wb, deuda_id):
    return next((d for d in _deuda_records(wb) if d['id'] == deuda_id), None)


def _save_deuda_status(wb, deuda_id, estado=None, proximo_vencimiento=None):
    ws = wb['Deudas Taller']
    row = find_row_by_id(ws, deuda_id)
    if not row:
        return
    headers = SHEET_HEADERS['Deudas Taller']
    if estado is not None:
        ws.cell(row=row[0].row, column=headers.index('Estado') + 1, value=estado)
    if proximo_vencimiento is not None:
        ws.cell(row=row[0].row, column=headers.index('Próximo Vencimiento') + 1, value=proximo_vencimiento)


@app.route('/api/deudas-taller', methods=['GET'])
def list_deudas_taller():
    return jsonify(_deudas_panel(get_wb(SHEETS_DEUDAS)))


@app.route('/api/deudas-taller/panel', methods=['GET'])
def deudas_taller_panel():
    return jsonify(_deudas_panel(get_wb(SHEETS_DEUDAS)))


@app.route('/api/deuda-taller', methods=['POST'])
def crear_deuda_taller():
    data = request.json or {}
    acreedor = ' '.join(str(data.get('acreedor') or '').strip().split())
    concepto = ' '.join(str(data.get('concepto') or '').strip().split())
    try:
        monto = round(float(data.get('monto_total') or 0), 2)
        frecuencia_base = _frequency(data.get('frecuencia'))
        tipo = _debt_type(data.get('tipo'), frecuencia_base)
        frecuencia = frecuencia_base if tipo == 'Recurrente' else 'Único'
    except (TypeError, ValueError) as exc:
        return jsonify({'success': False, 'error': str(exc) or 'Monto, tipo o frecuencia no válidos'}), 400
    if not acreedor or not concepto or monto <= 0:
        return jsonify({'success': False, 'error': 'Acreedor, concepto y un monto válido son obligatorios'}), 400
    registro = date.today()
    try:
        if tipo == 'Recurrente' and frecuencia == 'Mensual':
            dia = int(data.get('dia_pago') or 0)
            if dia < 1 or dia > 31:
                raise ValueError('El día de pago mensual debe estar entre 1 y 31')
            vencimiento = _monthly_due(dia, registro)
        else:
            vencimiento = _parse_iso_date(
                data.get('fecha_vencimiento'),
                'La primera fecha de pago semanal' if frecuencia == 'Semanal' else 'La fecha de vencimiento',
            )
            dia = vencimiento.day
    except (TypeError, ValueError) as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400

    wb = get_wb(SHEETS_DEUDAS)
    ws = wb['Deudas Taller']
    deuda_id = next_id(ws)
    append_row(ws, SHEET_HEADERS['Deudas Taller'], {
        'ID': deuda_id, 'Fecha Registro': registro.isoformat(), 'Acreedor': acreedor,
        'Concepto': concepto, 'Monto Total': monto, 'Frecuencia': frecuencia,
        'Día Pago': dia, 'Próximo Vencimiento': vencimiento.isoformat(),
        'Estado': 'Pendiente de reunir', 'Observaciones': (data.get('observaciones') or '').strip(),
        'Tipo': tipo,
    }, deuda_id)
    wb.save(EXCEL_FILE)
    return jsonify({'success': True, 'id': deuda_id, 'proximo_vencimiento': vencimiento.isoformat(),
                    'panel': _deudas_panel(wb)})


@app.route('/api/deuda-taller/<int:deuda_id>', methods=['PUT', 'PATCH'])
def update_deuda_taller(deuda_id):
    data = request.json or {}
    wb = get_wb(SHEETS_DEUDAS)
    actual = _get_deuda_summary(wb, deuda_id)
    if not actual:
        return jsonify({'success': False, 'error': 'Deuda no encontrada'}), 404

    acreedor = ' '.join(str(data.get('acreedor', actual['acreedor']) or '').strip().split())
    concepto = ' '.join(str(data.get('concepto', actual['concepto']) or '').strip().split())
    try:
        monto = round(float(data.get('monto_total', actual['monto_total']) or 0), 2)
        frecuencia_base = _frequency(data.get('frecuencia', actual['frecuencia']))
        tipo = _debt_type(data.get('tipo', actual['tipo']), frecuencia_base)
        frecuencia = frecuencia_base if tipo == 'Recurrente' else 'Único'
    except (TypeError, ValueError) as exc:
        return jsonify({'success': False, 'error': str(exc) or 'Monto, tipo o frecuencia no válidos'}), 400
    if not acreedor or not concepto or monto <= 0:
        return jsonify({'success': False, 'error': 'Acreedor, concepto y un monto válido son obligatorios'}), 400

    fondos = _fondos_deuda(wb)
    pagos = _pagos_deuda(wb)
    tiene_historial = any(x['deuda_id'] == deuda_id for x in fondos + pagos)
    if tiene_historial and (tipo != actual['tipo'] or frecuencia != actual['frecuencia']):
        return jsonify({'success': False, 'error': 'No se puede cambiar el tipo o la frecuencia de una deuda con fondos o pagos históricos'}), 409

    try:
        if tipo == 'Recurrente' and frecuencia == 'Mensual':
            dia = int(data.get('dia_pago', actual['dia_pago']) or 0)
            if dia < 1 or dia > 31:
                raise ValueError('El día de pago mensual debe estar entre 1 y 31')
            vencimiento = _monthly_due(dia, date.today()) if ('dia_pago' in data or actual['frecuencia'] != 'Mensual') else _parse_iso_date(actual['proximo_vencimiento'], 'Próximo vencimiento')
        else:
            raw_date = data.get('fecha_vencimiento', actual['proximo_vencimiento'])
            vencimiento = _parse_iso_date(raw_date, 'La primera fecha semanal' if frecuencia == 'Semanal' else 'La fecha de vencimiento')
            dia = vencimiento.day
    except (TypeError, ValueError) as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400

    periodo_actual = actual['periodo']
    pagos_periodo = sum(p['monto'] for p in pagos if p['deuda_id'] == deuda_id and (actual['frecuencia'] == 'Único' or p['periodo'] == periodo_actual))
    if monto + 0.009 < pagos_periodo:
        return jsonify({'success': False, 'error': f'El nuevo monto no puede ser menor que lo pagado en el período ({pagos_periodo:.2f})'}), 400

    ws = wb['Deudas Taller']
    row = find_row_by_id(ws, deuda_id)
    headers = SHEET_HEADERS['Deudas Taller']
    values = {
        'Acreedor': acreedor, 'Concepto': concepto, 'Monto Total': monto,
        'Frecuencia': frecuencia, 'Día Pago': dia, 'Próximo Vencimiento': vencimiento.isoformat(),
        'Estado': 'Pendiente de reunir', 'Observaciones': str(data.get('observaciones', actual['observaciones']) or '').strip(),
        'Tipo': tipo,
    }
    for key, value in values.items():
        ws.cell(row=row[0].row, column=headers.index(key) + 1, value=value)
    wb.save(EXCEL_FILE)
    return jsonify({'success': True, 'id': deuda_id, 'panel': _deudas_panel(wb)})


@app.route('/api/deuda-taller/<int:deuda_id>', methods=['DELETE'])
def delete_deuda_taller(deuda_id):
    wb = get_wb(SHEETS_DEUDAS)
    actual = _get_deuda_summary(wb, deuda_id)
    if not actual:
        return jsonify({'success': False, 'error': 'Deuda no encontrada'}), 404
    if any(x['deuda_id'] == deuda_id for x in _fondos_deuda(wb) + _pagos_deuda(wb)):
        return jsonify({'success': False, 'error': 'No se puede eliminar una deuda que ya tiene fondos o pagos registrados; edítala o déjala cerrada'}), 409
    ws = wb['Deudas Taller']
    row = find_row_by_id(ws, deuda_id)
    if not row:
        return jsonify({'success': False, 'error': 'Deuda no encontrada'}), 404
    ws.delete_rows(row[0].row, 1)
    wb.save(EXCEL_FILE)
    return jsonify({'success': True, 'id': deuda_id, 'panel': _deudas_panel(wb)})


@app.route('/api/fondo-deuda', methods=['POST'])
def create_fondo_deuda():
    data = request.json or {}
    try:
        deuda_id = int(data.get('deuda_id') or 0)
        monto = round(float(data.get('monto') or 0), 2)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'La deuda y el monto deben ser válidos'}), 400
    if deuda_id <= 0 or monto <= 0:
        return jsonify({'success': False, 'error': 'La deuda y el monto deben ser mayores que cero'}), 400
    wb = get_wb(SHEETS_DEUDAS)
    deuda = _get_deuda_summary(wb, deuda_id)
    if not deuda:
        return jsonify({'success': False, 'error': 'Deuda no encontrada'}), 404
    if monto > deuda['faltante_reunir'] + 0.009:
        return jsonify({'success': False, 'error': f'El aporte supera el faltante por reunir ({deuda["faltante_reunir"]:.2f})'}), 400
    ws = wb['Fondos Deudas']
    fondo_id = next_id(ws)
    fecha = data.get('fecha') or date.today().isoformat()
    try:
        fecha = _parse_iso_date(fecha, 'La fecha del aporte').isoformat()
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    append_row(ws, SHEET_HEADERS['Fondos Deudas'], {
        'ID': fondo_id, 'Fecha Aporte': fecha, 'Deuda ID': deuda_id,
        'Período': deuda['periodo'], 'Acreedor': deuda['acreedor'], 'Monto': monto,
        'Método': (data.get('metodo') or '').strip(), 'Observaciones': (data.get('observaciones') or '').strip(),
    }, fondo_id)
    wb.save(EXCEL_FILE)
    return jsonify({'success': True, 'id': fondo_id, 'panel': _deudas_panel(wb)})


@app.route('/api/pago-deuda', methods=['POST'])
def create_pago_deuda():
    data = request.json or {}
    try:
        deuda_id = int(data.get('deuda_id') or 0)
        monto = round(float(data.get('monto') or 0), 2)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'La deuda y el monto deben ser válidos'}), 400
    if deuda_id <= 0 or monto <= 0:
        return jsonify({'success': False, 'error': 'La deuda y el monto deben ser mayores que cero'}), 400
    wb = get_wb(SHEETS_DEUDAS)
    deuda = _get_deuda_summary(wb, deuda_id)
    if not deuda:
        return jsonify({'success': False, 'error': 'Deuda no encontrada'}), 404
    if monto > deuda['saldo_pendiente'] + 0.009:
        return jsonify({'success': False, 'error': f'El pago supera el saldo pendiente ({deuda["saldo_pendiente"]:.2f})'}), 400
    if monto > deuda['fondo_disponible'] + 0.009:
        return jsonify({'success': False, 'error': f'El bolsillo disponible no alcanza ({deuda["fondo_disponible"]:.2f})'}), 400
    tipo_pago = 'Total' if monto >= deuda['saldo_pendiente'] - 0.009 else 'Parcial'
    ws = wb['Pagos Deudas']
    pago_id = next_id(ws)
    fecha = data.get('fecha') or date.today().isoformat()
    try:
        fecha = _parse_iso_date(fecha, 'La fecha del pago').isoformat()
    except ValueError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 400
    append_row(ws, SHEET_HEADERS['Pagos Deudas'], {
        'ID': pago_id, 'Fecha Pago': fecha, 'Deuda ID': deuda_id,
        'Período': deuda['periodo'], 'Acreedor': deuda['acreedor'], 'Monto': monto,
        'Tipo Pago': tipo_pago, 'Observaciones': (data.get('observaciones') or '').strip(),
    }, pago_id)

    if tipo_pago == 'Total':
        current_due = _parse_iso_date(deuda['proximo_vencimiento'], 'Próximo vencimiento')
        if deuda['frecuencia'] == 'Mensual':
            next_due = _next_month_day(current_due, deuda['dia_pago'])
            _save_deuda_status(wb, deuda_id, 'Pendiente de reunir', next_due.isoformat())
        elif deuda['frecuencia'] == 'Semanal':
            next_due = _next_week_day(current_due)
            _save_deuda_status(wb, deuda_id, 'Pendiente de reunir', next_due.isoformat())
        else:
            _save_deuda_status(wb, deuda_id, 'Pagado')
    else:
        _save_deuda_status(wb, deuda_id, 'Pendiente de reunir')
    wb.save(EXCEL_FILE)
    return jsonify({'success': True, 'id': pago_id, 'tipo_pago': tipo_pago,
                    'panel': _deudas_panel(wb)})


@app.route('/api/nomina/descuento', methods=['POST'])
def aplicar_descuento_nomina():
    data = request.json or {}
    mecanico = ' '.join(str(data.get('mecanico') or '').strip().split())
    semana = str(data.get('semana') or '').strip()
    concepto = ' '.join(str(data.get('concepto') or '').strip().split())
    try:
        monto = round(float(data.get('monto') or 0), 2)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'El descuento debe ser numérico'}), 400
    if not mecanico or not semana or not concepto or monto <= 0:
        return jsonify({'success': False, 'error': 'Mecánico, semana, concepto y monto son obligatorios'}), 400

    raw_loan = data.get('prestamo_id')
    loan_id = None
    if raw_loan not in (None, '', 0, '0'):
        try:
            loan_id = int(raw_loan)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'El préstamo seleccionado no es válido'}), 400

    wb = get_wb()
    loans = _loan_records(wb)
    loan = next((p for p in loans if p['id'] == loan_id), None) if loan_id else None
    if loan_id and not loan:
        return jsonify({'success': False, 'error': 'Préstamo no encontrado'}), 404
    if loan and loan['mecanico'] != mecanico:
        return jsonify({'success': False, 'error': 'El préstamo no pertenece al mecánico seleccionado'}), 400
    if loan and monto > loan['saldo_pendiente']:
        return jsonify({'success': False, 'error': 'El descuento supera el saldo pendiente del préstamo'}), 400

    nomina = next((r for r in _nomina_calculada(wb, semana=semana, mecanico=mecanico)), None)
    bruto = nomina['bruto_pendiente'] if nomina else 0.0
    descuentos = _discount_records(wb)
    ya_aplicado = sum(d['monto'] for d in descuentos if d['mecanico'] == mecanico and d['semana'] == semana)
    disponible = max(round(bruto - ya_aplicado, 2), 0.0)
    if monto > disponible:
        return jsonify({'success': False, 'error': f'El descuento supera el saldo de nómina disponible ({disponible:.2f})'}), 400

    ws_d = wb['Descuentos Nomina']
    did = next_id(ws_d)
    fecha = data.get('fecha') or datetime.now().strftime('%Y-%m-%d')
    append_row(ws_d, SHEET_HEADERS['Descuentos Nomina'], {
        'ID': did, 'Fecha Aplicación': fecha, 'Mecánico': mecanico,
        'Semana': semana, 'Concepto': concepto, 'Monto': monto,
        'Préstamo ID': loan_id or '', 'Observaciones': (data.get('observaciones') or '').strip(),
    }, did)

    if loan:
        ws_l = wb['Prestamos']
        row_l = find_row_by_id(ws_l, loan_id)
        headers_l = SHEET_HEADERS['Prestamos']
        total_descontado = round(loan['total_descontado'] + monto, 2)
        saldo = max(round(loan['monto_original'] - total_descontado, 2), 0.0)
        ws_l.cell(row=row_l[0].row, column=headers_l.index('Total Descontado') + 1, value=total_descontado)
        ws_l.cell(row=row_l[0].row, column=headers_l.index('Saldo Pendiente') + 1, value=saldo)
        ws_l.cell(row=row_l[0].row, column=headers_l.index('Estado') + 1, value='Pagado' if saldo <= 0.009 else 'Pendiente')

    wb.save(EXCEL_FILE)
    return jsonify({'success': True, 'id': did, 'monto': monto, 'saldo_prestamo': loan['saldo_pendiente'] - monto if loan else None,
                    'nomina': _nomina_response(wb, semana=semana)})


@app.route('/api/nomina/pagar', methods=['POST'])
def pagar_nomina():
    data = request.json
    mecanico = data.get("mecanico")
    semana = data.get("semana")
    if not mecanico or not semana:
        return jsonify({"success": False, "error": "Falta mecánico o semana"}), 400

    wb = get_wb()
    ws = wb["Trabajos"]
    headers = SHEET_HEADERS["Trabajos"]
    fecha_pago = datetime.now().strftime("%Y-%m-%d")
    n_trabajos, total_mo, total_comision = 0, 0.0, 0.0

    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        vals = {headers[i]: row[i].value for i in range(len(headers))}
        if (vals.get("Mecánico") == mecanico and vals.get("Semana") == semana
                and vals.get("Estado") != "Pagado" and vals.get("Factura Cancelada") == "Sí"):
            n_trabajos += 1
            total_mo += float(vals.get("Monto Mano de Obra") or 0)
            total_comision += float(vals.get("Monto Mecánico") or 0)
            row[headers.index("Estado")].value = "Pagado"
            row[headers.index("Fecha Pago")].value = fecha_pago

    descuentos_periodo = sum(
        d['monto'] for d in _discount_records(wb)
        if d['mecanico'] == mecanico and d['semana'] == semana
    )
    # Si existía un pago anterior de la misma semana, sus descuentos ya fueron
    # aplicados y no deben restarse otra vez sobre trabajos nuevos.
    pagos_previos = 0.0
    ws_pagos = wb['Pagos']
    headers_pagos = SHEET_HEADERS['Pagos']
    for row_pago in ws_pagos.iter_rows(min_row=2):
        if row_pago[0].value is None:
            continue
        vals_pago = _row_values(row_pago, headers_pagos)
        if vals_pago.get('Mecánico') == mecanico and vals_pago.get('Semana') == semana:
            pagos_previos += money(vals_pago.get('Total Descuentos'))
    total_descuentos = round(max(descuentos_periodo - pagos_previos, 0.0), 2)
    total_descuentos = min(total_descuentos, round(total_comision, 2))
    neto_pagado = round(max(total_comision - total_descuentos, 0), 2)

    if n_trabajos > 0:
        ws_pagos = wb["Pagos"]
        pid = next_id(ws_pagos)
        append_row(ws_pagos, SHEET_HEADERS["Pagos"], {
            "ID": pid, "Fecha Pago": fecha_pago, "Mecánico": mecanico,
            "Semana": semana, "N° Trabajos": n_trabajos,
            "Total Mano de Obra": round(total_mo, 2), "Total Comisión": round(total_comision, 2),
            "Total Descuentos": total_descuentos, "Neto Pagado": neto_pagado,
        }, pid)

    wb.save(EXCEL_FILE)
    return jsonify({"success": True, "n_trabajos": n_trabajos,
                    "total_comision": round(total_comision, 2),
                    "total_descuentos": total_descuentos,
                    "neto_pagado": neto_pagado})


@app.route('/api/pagos', methods=['GET'])
def list_pagos():
    wb = get_wb()
    ws = wb["Pagos"]
    data = sheet_to_dicts(ws, SHEET_HEADERS["Pagos"])
    data.sort(key=lambda d: d.get("ID") or 0, reverse=True)
    return jsonify(data)


# ---------------------------------------------------------------------------
# API Gastos
# ---------------------------------------------------------------------------

@app.route('/api/gastos', methods=['GET'])
def list_gastos():
    wb = get_wb()
    ws = wb["Gastos"]
    data = sheet_to_dicts(ws, SHEET_HEADERS["Gastos"])
    fecha = request.args.get('fecha')
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')
    categoria = request.args.get('categoria')
    if fecha:
        data = [d for d in data if d.get("Fecha") == fecha]
    if desde:
        data = [d for d in data if (d.get("Fecha") or "") >= desde]
    if hasta:
        data = [d for d in data if (d.get("Fecha") or "") <= hasta]
    if categoria:
        data = [d for d in data if d.get("Categoría") == categoria]
    data.sort(key=lambda d: (d.get("Fecha") or "", d.get("ID") or 0), reverse=True)
    return jsonify(data)


@app.route('/api/gasto', methods=['POST'])
def create_gasto():
    data = request.json
    wb = get_wb()
    ws = wb["Gastos"]
    gid = next_id(ws)
    append_row(ws, SHEET_HEADERS["Gastos"], {
        "ID": gid,
        "Fecha": data.get("fecha", datetime.now().strftime("%Y-%m-%d")),
        "Categoría": data.get("categoria", "Otros"),
        "Descripción": data.get("descripcion", ""),
        "Monto": float(data.get("monto", 0) or 0),
        "Responsable": data.get("responsable", ""),
        "Método de Pago": data.get("metodo_pago", ""),
    }, gid)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True, "id": gid})


@app.route('/api/gasto/<int:gid>', methods=['DELETE'])
def delete_gasto(gid):
    wb = get_wb()
    ws = wb["Gastos"]
    row = find_row_by_id(ws, gid)
    if row:
        ws.delete_rows(row[0].row, 1)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/gastos/resumen', methods=['GET'])
def resumen_gastos():
    fecha = request.args.get('fecha', datetime.now().strftime("%Y-%m-%d"))
    wb = get_wb()
    ws = wb["Gastos"]
    data = sheet_to_dicts(ws, SHEET_HEADERS["Gastos"])

    hoy = [d for d in data if d.get("Fecha") == fecha]
    total_hoy = sum(float(d.get("Monto") or 0) for d in hoy)
    por_categoria = {}
    for d in hoy:
        cat = d.get("Categoría", "Otros")
        por_categoria[cat] = por_categoria.get(cat, 0) + float(d.get("Monto") or 0)

    mes = fecha[:7]
    del_mes = [d for d in data if (d.get("Fecha") or "").startswith(mes)]
    total_mes = sum(float(d.get("Monto") or 0) for d in del_mes)

    return jsonify({
        "fecha": fecha, "total_hoy": round(total_hoy, 2),
        "por_categoria": por_categoria, "total_mes": round(total_mes, 2),
    })


# ---------------------------------------------------------------------------
# API Herramientas
# ---------------------------------------------------------------------------

@app.route('/api/herramientas', methods=['GET'])
def list_herramientas():
    wb = get_wb()
    ws = wb["Herramientas"]
    data = sheet_to_dicts(ws, SHEET_HEADERS["Herramientas"])
    estado = request.args.get('estado')
    if estado:
        data = [d for d in data if d.get("Estado") == estado]
    data.sort(key=lambda d: (d.get("Fecha Préstamo") or "", d.get("ID") or 0), reverse=True)
    return jsonify(data)


@app.route('/api/herramienta', methods=['POST'])
def create_herramienta():
    data = request.json
    wb = get_wb()
    ws = wb["Herramientas"]
    hid = next_id(ws)
    append_row(ws, SHEET_HEADERS["Herramientas"], {
        "ID": hid,
        "Herramienta": (data.get("herramienta") or "").strip(),
        "Prestada A": (data.get("prestada_a") or "").strip(),
        "Entregada Por": (data.get("entregada_por") or "").strip(),
        "Fecha Préstamo": data.get("fecha_prestamo", datetime.now().strftime("%Y-%m-%d")),
        "Fecha Devolución": "",
        "Estado": "Prestada",
        "Observaciones": data.get("observaciones", ""),
    }, hid)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True, "id": hid})


@app.route('/api/herramienta/<int:hid>', methods=['PUT'])
def update_herramienta(hid):
    data = request.json
    wb = get_wb()
    ws = wb["Herramientas"]
    row = find_row_by_id(ws, hid)
    if not row:
        return jsonify({"success": False, "error": "No encontrado"}), 404
    headers = SHEET_HEADERS["Herramientas"]

    def set_val(header, value):
        ws.cell(row=row[0].row, column=headers.index(header) + 1, value=value)

    if data.get("devolver"):
        set_val("Fecha Devolución", datetime.now().strftime("%Y-%m-%d"))
        set_val("Estado", "Devuelta")
    field_map = {"herramienta": "Herramienta", "prestada_a": "Prestada A",
                 "entregada_por": "Entregada Por", "observaciones": "Observaciones"}
    for key, header in field_map.items():
        if key in data:
            set_val(header, data[key])
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


@app.route('/api/herramienta/<int:hid>', methods=['DELETE'])
def delete_herramienta(hid):
    wb = get_wb()
    ws = wb["Herramientas"]
    row = find_row_by_id(ws, hid)
    if row:
        ws.delete_rows(row[0].row, 1)
    wb.save(EXCEL_FILE)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@app.route('/api/excel', methods=['GET'])
def download_excel():
    content = workbook_to_bytes(EXCEL_FILE)
    return send_file(io.BytesIO(content),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='Control_Taller_VeneAutos.xlsx')


# Vercel carga este módulo desde el WSGI unificado de ../app.py.
