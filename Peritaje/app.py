from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.exceptions import HTTPException
import json, os, base64
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from storage import load_workbook_for_app, workbook_to_bytes
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
import io

app = Flask(__name__,
            template_folder=os.path.join(os.path.dirname(__file__), 'templates'),
            static_folder=os.path.join(os.path.dirname(__file__), 'static'),
            static_url_path='/static')

@app.errorhandler(Exception)
def api_error(error):
    if isinstance(error, HTTPException):
        return error
    app.logger.exception("Error no controlado en Peritaje")
    return jsonify({"success": False, "error": f"{type(error).__name__}: {error}"}), 500

EXCEL_FILE = "google-sheets://Peritajes"
LOGO_PATH = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')

ITEMS_PERITAJE = [
    "Motor", "Caja de cambios", "Frenos delanteros", "Frenos traseros",
    "Suspensión delantera", "Suspensión trasera", "Dirección", "Sistema eléctrico",
    "Batería", "Alternador", "Sistema de enfriamiento", "Aire acondicionado",
    "Llantas", "Aros", "Parabrisas", "Lunas laterales", "Luces delanteras",
    "Luces traseras", "Carrocería", "Pintura", "Interior / Tapicería",
    "Tablero / Instrumentos", "Transmisión", "Escape / Silenciador",
    "Filtros (aire, aceite, combustible)", "Correa de distribución", "Amortiguadores"
]

ESTADOS = ["Bueno", "Regular", "Malo", "No aplica"]

def get_next_number():
    wb = load_workbook_for_app(EXCEL_FILE, data_only=True, read_only=True)
    try:
        ws = wb['Peritajes'] if 'Peritajes' in wb.sheetnames else wb.active
        return max((ws.max_row or 1), 1)  # encabezado ocupa la fila 1
    finally:
        wb.close()

def save_peritaje(data):
    wb = load_workbook_for_app(EXCEL_FILE)
    ws = wb['Peritajes'] if 'Peritajes' in wb.sheetnames else wb.create_sheet('Peritajes')

    items = data.get("items", {})
    # Usar los ítems que vienen del formulario (incluye los custom)
    item_keys = list(items.keys())

    base_headers = ["N° Peritaje", "Fecha", "N° OT", "Cliente", "Placa",
                    "Marca", "Modelo", "Año", "Color", "Kilometraje",
                    "Técnico", "Observaciones Generales", "Total Estimado"]

    # Leer encabezados existentes o crear desde cero
    existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if not existing_headers[0]:  # hoja vacía
        existing_headers = []

    # Agregar columnas nuevas que no existan todavía
    all_headers = list(existing_headers) if existing_headers else base_headers[:]
    for key in item_keys:
        if key not in all_headers:
            all_headers.append(key)

    # Reescribir fila de encabezados si cambió
    if all_headers != existing_headers:
        for col, h in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="C0392B")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
        ws.row_dimensions[1].height = 35
        for col in range(1, len(all_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14

    num = ws.max_row  # fila 1 = header, siguiente = num
    row_num = ws.max_row + 1
    num_peritaje = f"PER-{num:04d}"

    row_values = {
        "N° Peritaje": num_peritaje,
        "Fecha": data.get("fecha", datetime.now().strftime("%d/%m/%Y")),
        "N° OT": data.get("num_ot", ""),
        "Cliente": data.get("cliente", ""),
        "Placa": data.get("placa", "").upper(),
        "Marca": data.get("marca", ""),
        "Modelo": data.get("modelo", ""),
        "Año": data.get("anio", ""),
        "Color": data.get("color", ""),
        "Kilometraje": data.get("kilometraje", ""),
        "Técnico": data.get("tecnico", ""),
        "Observaciones Generales": data.get("observaciones", ""),
        "Total Estimado": data.get("total_estimado", 0),
    }
    for item_name, item_data in items.items():
        row_values[item_name] = item_data.get("estado", "")

    fill_color = "F9EBEA" if row_num % 2 == 0 else "FFFFFF"
    for col, header in enumerate(all_headers, 1):
        value = row_values.get(header, "")
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        if col == 1:
            cell.font = Font(bold=True, color="C0392B")

    wb.save(EXCEL_FILE)
    return num_peritaje

def get_all_peritajes():
    wb = load_workbook_for_app(EXCEL_FILE, data_only=True, read_only=True)
    try:
        ws = wb['Peritajes'] if 'Peritajes' in wb.sheetnames else wb.active
        headers = [cell.value for cell in ws[1]] if ws.max_column else []
        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                result.append(dict(zip(headers, row)))
        return result
    finally:
        wb.close()

def generate_pdf(peritaje_data, num_peritaje):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Header with logo
    logo_img = None
    if os.path.exists(LOGO_PATH):
        logo_img = Image(LOGO_PATH, width=4*cm, height=3*cm)
    
    title_style = ParagraphStyle('title', fontSize=14, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1a1a1a'), alignment=TA_CENTER)
    sub_style = ParagraphStyle('sub', fontSize=9, fontName='Helvetica',
        textColor=colors.HexColor('#555555'), alignment=TA_CENTER)
    
    header_data = [[
        logo_img if logo_img else '',
        [Paragraph("VENE AUTOS - TALLER AUTOMOTRIZ", title_style),
         Paragraph("Tel. 3225167224 · veneautos82@gmail.com", sub_style),
         Paragraph("Lunes a Sábado 8am - 6pm", sub_style)]
    ]]
    header_table = Table(header_data, colWidths=[5*cm, 13*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
    ]))
    story.append(header_table)
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#C0392B')))
    story.append(Spacer(1, 0.3*cm))
    
    # OT Info bar
    ot_style = ParagraphStyle('ot', fontSize=12, fontName='Helvetica-Bold',
        textColor=colors.white, alignment=TA_LEFT)
    date_style = ParagraphStyle('date', fontSize=9, fontName='Helvetica',
        textColor=colors.white, alignment=TA_RIGHT)
    
    ot_bar = Table([[
        Paragraph(f"PERITAJE DE VEHÍCULO · {num_peritaje}", ot_style),
        Paragraph(f"Fecha: {peritaje_data.get('fecha','')}<br/>OT: {peritaje_data.get('num_ot','N/A')}", date_style)
    ]], colWidths=[10*cm, 8*cm])
    ot_bar.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#C0392B')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (0,0), 8),
        ('RIGHTPADDING', (-1,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(ot_bar)
    story.append(Spacer(1, 0.4*cm))
    
    # Vehicle & Client info
    lbl = ParagraphStyle('lbl', fontSize=8, fontName='Helvetica-Bold', textColor=colors.HexColor('#777777'))
    val = ParagraphStyle('val', fontSize=10, fontName='Helvetica-Bold', textColor=colors.HexColor('#1a1a1a'))
    
    def field(label, value):
        return [Paragraph(label, lbl), Paragraph(str(value or '—'), val)]
    
    info_data = [
        [field("CLIENTE", peritaje_data.get('cliente','')),
         field("PLACA", peritaje_data.get('placa',''))],
        [field("VEHÍCULO", f"{peritaje_data.get('marca','')} {peritaje_data.get('modelo','')}"),
         field("AÑO", peritaje_data.get('anio',''))],
        [field("COLOR", peritaje_data.get('color','')),
         field("KILOMETRAJE", f"{peritaje_data.get('kilometraje','')} km")],
        [field("TÉCNICO", peritaje_data.get('tecnico','')),
         field("ESTADO GENERAL", peritaje_data.get('estado_general',''))],
    ]
    
    for row in info_data:
        t = Table([row], colWidths=[9*cm, 9*cm])
        t.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(t)
    
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#DDDDDD')))
    story.append(Spacer(1, 0.3*cm))
    
    # Items section title
    sec_style = ParagraphStyle('sec', fontSize=11, fontName='Helvetica-Bold',
        textColor=colors.white, alignment=TA_LEFT)
    sec_bar = Table([[Paragraph("INSPECCIÓN DE COMPONENTES", sec_style)]],
        colWidths=[18*cm])
    sec_bar.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#2C3E50')),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(sec_bar)
    story.append(Spacer(1, 0.2*cm))
    
    # Items grid
    items = peritaje_data.get('items', {})
    h_style = ParagraphStyle('h', fontSize=8, fontName='Helvetica-Bold',
        textColor=colors.white, alignment=TA_CENTER)
    item_name_style = ParagraphStyle('in', fontSize=8, fontName='Helvetica',
        textColor=colors.HexColor('#1a1a1a'))
    
    estado_colors = {
        'Bueno': '#27AE60',
        'Regular': '#F39C12',
        'Malo': '#E74C3C',
        'No aplica': '#95A5A6'
    }
    
    grid_data = [[
        Paragraph("COMPONENTE", h_style),
        Paragraph("ESTADO", h_style),
        Paragraph("OBSERVACIÓN", h_style),
        Paragraph("COMPONENTE", h_style),
        Paragraph("ESTADO", h_style),
        Paragraph("OBSERVACIÓN", h_style),
    ]]
    
    # Usar los ítems que vienen del formulario (incluye custom)
    item_list = list(items.keys()) if items else list(ITEMS_PERITAJE)
    for i in range(0, len(item_list), 2):
        row = []
        for j in range(2):
            if i+j < len(item_list):
                item_name = item_list[i+j]
                item_data = items.get(item_name, {})
                estado = item_data.get('estado', '')
                obs = item_data.get('obs', '')
                row.append(Paragraph(item_name, item_name_style))
                ec = estado_colors.get(estado, '#CCCCCC')
                estado_p = ParagraphStyle('ep', fontSize=8, fontName='Helvetica-Bold',
                    textColor=colors.white, alignment=TA_CENTER,
                    backColor=colors.HexColor(ec))
                row.append(Paragraph(estado or '', estado_p))
                row.append(Paragraph(obs or '', item_name_style))
            else:
                row += ['', '', '']
        grid_data.append(row)
    
    col_w = [4.5*cm, 2.2*cm, 2.3*cm, 4.5*cm, 2.2*cm, 2.3*cm]
    items_table = Table(grid_data, colWidths=col_w, repeatRows=1)
    
    ts = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F8F9FA'), colors.white]),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
    ])
    
    # Color estado cells
    for r_idx, row in enumerate(grid_data[1:], 1):
        for c_idx in [1, 4]:
            if c_idx < len(row):
                cell_val = ''
                if hasattr(row[c_idx], 'text'):
                    cell_val = row[c_idx].text if hasattr(row[c_idx], 'text') else ''
                estado = grid_data[r_idx][c_idx]
                if hasattr(estado, 'style') and hasattr(estado.style, 'backColor'):
                    ts.add('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), estado.style.backColor)
    
    items_table.setStyle(ts)
    story.append(items_table)
    story.append(Spacer(1, 0.4*cm))
    
    # Observaciones
    obs_gen = peritaje_data.get('observaciones', '')
    total = peritaje_data.get('total_estimado', 0)
    
    if obs_gen:
        obs_title = Table([[Paragraph("OBSERVACIONES GENERALES", sec_style)]],
            colWidths=[18*cm])
        obs_title.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#2C3E50')),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(obs_title)
        obs_p = ParagraphStyle('obsp', fontSize=9, fontName='Helvetica',
            textColor=colors.HexColor('#333333'), leading=14)
        obs_box = Table([[Paragraph(obs_gen, obs_p)]], colWidths=[18*cm])
        obs_box.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(Spacer(1, 0.2*cm))
        story.append(obs_box)
        story.append(Spacer(1, 0.3*cm))
    
    # Total estimado
    if total:
        total_style = ParagraphStyle('tot', fontSize=11, fontName='Helvetica-Bold',
            textColor=colors.white, alignment=TA_RIGHT)
        total_table = Table([[
            Paragraph(f"TOTAL ESTIMADO REPARACIÓN:  ${int(total):,}".replace(',', '.'), total_style)
        ]], colWidths=[18*cm])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#C0392B')),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 7),
            ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ]))
        story.append(total_table)
        story.append(Spacer(1, 0.5*cm))
    
    # Firmas
    firma_style = ParagraphStyle('firma', fontSize=9, fontName='Helvetica', alignment=TA_CENTER)
    firma_label = ParagraphStyle('firmalbl', fontSize=8, fontName='Helvetica',
        textColor=colors.HexColor('#777777'), alignment=TA_CENTER)
    
    firmas = Table([
        [Paragraph("_______________________________", firma_style),
         Paragraph("_______________________________", firma_style)],
        [Paragraph("Firma del Cliente", firma_label),
         Paragraph("Firma del Técnico", firma_label)],
    ], colWidths=[9*cm, 9*cm])
    story.append(firmas)
    
    # Footer
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#DDDDDD')))
    footer_style = ParagraphStyle('footer', fontSize=7, fontName='Helvetica',
        textColor=colors.HexColor('#999999'), alignment=TA_CENTER)
    story.append(Paragraph(
        "DOCUMENTO NO FISCAL · Persona natural no obligada a facturar electrónicamente (Art. 437 E.T.)",
        footer_style))
    
    doc.build(story)
    buf.seek(0)
    return buf

@app.route('/')
def index():
    return render_template('index.html', items=ITEMS_PERITAJE, estados=ESTADOS)

@app.route('/api/peritajes', methods=['GET'])
def list_peritajes():
    return jsonify(get_all_peritajes())

@app.route('/api/peritaje', methods=['POST'])
def create_peritaje():
    data = request.json
    num = save_peritaje(data)
    return jsonify({"success": True, "num_peritaje": num})

@app.route('/api/peritaje/pdf', methods=['POST'])
def export_pdf():
    data = request.json
    num = data.get('num_peritaje', 'PREVIEW')
    buf = generate_pdf(data, num)
    return send_file(buf, mimetype='application/pdf',
        as_attachment=True, download_name=f"Peritaje_{num}.pdf")

@app.route('/api/excel', methods=['GET'])
def download_excel():
    content = workbook_to_bytes(EXCEL_FILE)
    return send_file(io.BytesIO(content), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='Peritajes_VeneAutos.xlsx')

# Vercel carga esta aplicación desde el WSGI unificado de ../app.py.
