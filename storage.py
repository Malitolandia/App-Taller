"""Persistencia exclusiva de APPTALLER en Google Sheets.

Vercel no usa el sistema de archivos como base de datos. Los módulos siguen
trabajando con objetos compatibles con openpyxl en memoria, pero
``RemoteWorkbook.save()`` sincroniza únicamente las pestañas modificadas con
una hoja de cálculo de Google.
"""
from __future__ import annotations

import io
import json
import os
import random
import threading
import time
from datetime import date, datetime
from functools import lru_cache
from typing import BinaryIO, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

SHEETS_ID_ENV = "GOOGLE_SHEETS_ID"
SERVICE_JSON_ENV = "GOOGLE_SERVICE_ACCOUNT_JSON"

CONTROL_HEADERS = {
    "Mecanicos": ["ID", "Nombre", "Teléfono", "% Comisión", "Activo", "Fecha Ingreso"],
    "Equipos": ["ID", "Nombre", "Integrante 1", "Integrante 2", "% Comisión Total", "Activo"],
    "Trabajos": [
        "ID", "Fecha", "Semana", "Mecánico", "Equipo", "Placa", "Vehículo",
        "Descripción", "Monto Mano de Obra", "% Comisión", "Monto Mecánico",
        "Factura Cancelada", "Estado", "Fecha Pago", "Grupo",
    ],
    "Gastos": ["ID", "Fecha", "Categoría", "Descripción", "Monto", "Responsable", "Método de Pago"],
    "Pagos": [
        "ID", "Fecha Pago", "Mecánico", "Semana", "N° Trabajos", "Total Mano de Obra",
        "Total Comisión", "Total Descuentos", "Neto Pagado",
    ],
    "Prestamos": [
        "ID", "Fecha", "Mecánico", "Monto Original", "Cuota Sugerida", "Total Descontado",
        "Saldo Pendiente", "Estado", "Observaciones",
    ],
    "Descuentos Nomina": [
        "ID", "Fecha Aplicación", "Mecánico", "Semana", "Concepto", "Monto",
        "Préstamo ID", "Observaciones",
    ],
    "Deudas Taller": [
        "ID", "Fecha Registro", "Acreedor", "Concepto", "Monto Total", "Frecuencia",
        "Día Pago", "Próximo Vencimiento", "Estado", "Observaciones",
    ],
    "Fondos Deudas": [
        "ID", "Fecha Aporte", "Deuda ID", "Período", "Acreedor", "Monto", "Método", "Observaciones",
    ],
    "Pagos Deudas": [
        "ID", "Fecha Pago", "Deuda ID", "Período", "Acreedor", "Monto", "Tipo Pago", "Observaciones",
    ],
    "Herramientas": [
        "ID", "Herramienta", "Prestada A", "Entregada Por", "Fecha Préstamo",
        "Fecha Devolución", "Estado", "Observaciones",
    ],
}

NEVERAS_HEADERS = {
    "Ventas": ["#", "Fecha", "Hora", "Cliente", "Producto", "Cantidad", "Precio Unit.", "Total", "Método Pago", "Pagó", "Estado Pago", "Ganancia"],
    "Inventario": ["Producto", "Costo", "Precio Venta", "Ganancia Unit.", "Stock Inicial", "Stock Mín.", "Vendidos", "Stock Actual", "Estado", "Costo Total Inv.", "Ganancia x Producto"],
    "Clientes": ["Cliente", "Total Comprado", "Total Pagado", "Deuda Pendiente", "N° Compras", "Última Compra", "Estado"],
    "Deudas": ["Cliente", "Monto"],
    "Dashboard": ["Indicador", "Valor"],
}

PERITAJE_BASE_HEADERS = [
    "N° Peritaje", "Fecha", "N° OT", "Cliente", "Placa", "Marca", "Modelo", "Año",
    "Color", "Kilometraje", "Técnico", "Observaciones Generales", "Total Estimado",
    "Motor", "Caja de cambios", "Frenos delanteros", "Frenos traseros",
    "Suspensión delantera", "Suspensión trasera", "Dirección", "Sistema eléctrico",
    "Batería", "Alternador", "Sistema de enfriamiento", "Aire acondicionado",
    "Llantas", "Aros", "Parabrisas", "Lunas laterales", "Luces delanteras",
    "Luces traseras", "Carrocería", "Pintura", "Interior / Tapicería",
    "Tablero / Instrumentos", "Transmisión", "Escape / Silenciador",
    "Filtros (aire, aceite, combustible)", "Correa de distribución", "Amortiguadores",
]

REMOTE_HEADERS = {
    **CONTROL_HEADERS,
    **NEVERAS_HEADERS,
    "Peritajes": PERITAJE_BASE_HEADERS,
}


def _credentials_info() -> dict | None:
    raw = os.getenv(SERVICE_JSON_ENV, "").strip()
    if raw:
        try:
            value = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON no contiene JSON válido") from exc
        if not isinstance(value, dict) or not value.get("client_email"):
            raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON no contiene una cuenta de servicio válida")
        return value

    return None


def sheets_enabled() -> bool:
    """Indica si el backend remoto está configurado; no existe fallback local."""
    return bool(os.getenv(SHEETS_ID_ENV, "").strip() and _credentials_info())


@lru_cache(maxsize=1)
def _sheets_service():
    info = _credentials_info()
    spreadsheet_id = os.getenv(SHEETS_ID_ENV, "").strip()
    if not spreadsheet_id:
        raise RuntimeError("Falta GOOGLE_SHEETS_ID en las variables de entorno")
    if not info:
        raise RuntimeError(
            "Faltan credenciales de Google. Configura GOOGLE_SERVICE_ACCOUNT_JSON en Vercel"
        )
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "Faltan google-api-python-client y google-auth en requirements.txt"
        ) from exc

    credentials = Credentials.from_service_account_info(
        info,
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def _spreadsheet_id() -> str:
    value = os.getenv(SHEETS_ID_ENV, "").strip()
    if not value:
        raise RuntimeError("Falta GOOGLE_SHEETS_ID en las variables de entorno")
    return value


def _quote_title(title: str) -> str:
    return "'" + title.replace("'", "''") + "'"


def _title_range(title: str) -> str:
    # La API acepta el título sin coordenadas para referirse a toda la pestaña.
    return _quote_title(title)


def google_error_status(error) -> int | None:
    """Mapea límites temporales de Google Sheets a una respuesta HTTP recuperable."""
    response = getattr(error, "resp", None)
    status = getattr(response, "status", None)
    text = str(error).upper()
    if status == 429 or "RATE_LIMIT_EXCEEDED" in text or "QUOTA EXCEEDED" in text:
        return 503
    return None


def _execute(request, retries: int = 3):
    """Ejecuta una llamada Google con backoff para 429/5xx."""
    for attempt in range(retries + 1):
        try:
            try:
                return request.execute(num_retries=0)
            except TypeError:
                return request.execute()
        except Exception as exc:
            response = getattr(exc, "resp", None)
            status = getattr(response, "status", None)
            if status not in {429, 500, 502, 503, 504} or attempt >= retries:
                raise
            delay = min(8.0, 0.75 * (2 ** attempt) + random.random() * 0.25)
            time.sleep(delay)
    raise RuntimeError("La llamada a Google Sheets no pudo completarse")


_PROPERTIES_CACHE: tuple[float, list[dict]] | None = None
_PROPERTIES_TTL_SECONDS = 20.0


def _sheet_properties(force_refresh: bool = False) -> list[dict]:
    global _PROPERTIES_CACHE
    now = time.monotonic()
    if not force_refresh and _PROPERTIES_CACHE and now - _PROPERTIES_CACHE[0] < _PROPERTIES_TTL_SECONDS:
        return [dict(item) for item in _PROPERTIES_CACHE[1]]
    response = _execute(
        _sheets_service()
        .spreadsheets()
        .get(
            spreadsheetId=_spreadsheet_id(),
            includeGridData=False,
            fields="sheets.properties",
        )
    )
    properties = [sheet["properties"] for sheet in response.get("sheets", [])]
    _PROPERTIES_CACHE = (now, properties)
    return [dict(item) for item in properties]


_SCHEMA_READY = False
_SCHEMA_LOCK = threading.Lock()


def _grid_requests(properties: list[dict]) -> list[dict]:
    by_title = {item.get("title"): item for item in properties}
    requests = []
    for title, headers in REMOTE_HEADERS.items():
        item = by_title.get(title)
        if not item or not item.get("sheetId"):
            continue
        grid = item.get("gridProperties", {})
        rows = int(grid.get("rowCount") or 1000)
        cols = int(grid.get("columnCount") or 26)
        target_rows = max(rows, 1000)
        # Peritajes puede conservar campos personalizados de respaldos antiguos.
        target_cols = max(cols, len(headers), 100)
        if target_rows != rows or target_cols != cols:
            requests.append({
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": item["sheetId"],
                        "gridProperties": {
                            "rowCount": target_rows,
                            "columnCount": target_cols,
                        },
                    },
                    "fields": "gridProperties(rowCount,columnCount)",
                }
            })
    return requests


def _ensure_required_sheets() -> None:
    """Crea las pestañas y encabezados sin borrar datos existentes."""
    service = _sheets_service()
    properties = _sheet_properties()
    existing = {item["title"] for item in properties}
    missing = [title for title in REMOTE_HEADERS if title not in existing]
    if missing:
        _execute(service.spreadsheets().batchUpdate(
            spreadsheetId=_spreadsheet_id(),
            body={
                "requests": [
                    {
                        "addSheet": {
                            "properties": {
                                "title": title,
                                "gridProperties": {
                                    "rowCount": 1000,
                                    "columnCount": max(100, len(REMOTE_HEADERS[title])),
                                },
                            }
                        }
                    }
                    for title in missing
                ]
            },
        ))
        properties = _sheet_properties(force_refresh=True)

    grid_requests = _grid_requests(properties)
    if grid_requests:
        _execute(service.spreadsheets().batchUpdate(
            spreadsheetId=_spreadsheet_id(),
            body={"requests": grid_requests},
        ))

    values_api = service.spreadsheets().values()
    titles = list(REMOTE_HEADERS)
    header_response = _execute(values_api.batchGet(
        spreadsheetId=_spreadsheet_id(),
        ranges=[
            f"{_quote_title(title)}!A1:{get_column_letter(max(len(REMOTE_HEADERS[title]), 1))}1"
            for title in titles
        ],
        valueRenderOption="UNFORMATTED_VALUE",
        dateTimeRenderOption="FORMATTED_STRING",
    ))
    value_ranges = header_response.get("valueRanges", [])
    for index, (title, headers) in enumerate(REMOTE_HEADERS.items()):
        value_range = value_ranges[index] if index < len(value_ranges) else {}
        first_row = value_range.get("values", [])
        populated = bool(first_row and any(value not in (None, "") for value in first_row[0]))
        # Pagos recibió columnas de descuentos y neto en esta versión. Solo se
        # amplía si la fila existente conserva el encabezado anterior; no se
        # reemplazan encabezados personalizados de las demás pestañas.
        needs_payments_upgrade = (
            title == "Pagos" and populated
            and list(first_row[0][:len(headers)]) != list(headers)
            and list(first_row[0][:7]) == ["ID", "Fecha Pago", "Mecánico", "Semana", "N° Trabajos", "Total Mano de Obra", "Total Comisión"]
        )
        if not populated or needs_payments_upgrade:
            response = _execute(values_api.update(
                spreadsheetId=_spreadsheet_id(),
                range=f"{_quote_title(title)}!A1",
                valueInputOption="RAW",
                body={"values": [headers]},
            ))
            if not response.get("updatedRange"):
                raise RuntimeError(f"Google Sheets no confirmó los encabezados de '{title}'")


def _prepare_remote_schema() -> None:
    global _SCHEMA_READY
    if _SCHEMA_READY:
        return
    with _SCHEMA_LOCK:
        if _SCHEMA_READY:
            return
        _ensure_required_sheets()
        _SCHEMA_READY = True


_VALUES_CACHE: tuple[float, dict[str, list[list]]] | None = None
# Caché breve por conjunto de pestañas. La clave ordenada permite reutilizar
# la instantánea exacta de Nómina o Deudas sin cargar las 17 pestañas.
_SCOPED_VALUES_CACHE: dict[tuple[str, ...], tuple[float, dict[str, list[list]]]] = {}
_VALUES_TTL_SECONDS = 10.0


def _invalidate_values_cache() -> None:
    global _VALUES_CACHE
    _VALUES_CACHE = None
    _SCOPED_VALUES_CACHE.clear()


def _scope_key(sheet_titles: Iterable[str] | None) -> tuple[str, ...] | None:
    if sheet_titles is None:
        return None
    wanted = {str(title).strip() for title in sheet_titles if str(title).strip()}
    return tuple(title for title in REMOTE_HEADERS if title in wanted)


def _cache_entry(scope: tuple[str, ...] | None):
    if scope is None:
        return _VALUES_CACHE
    return _SCOPED_VALUES_CACHE.get(scope)


def _cache_is_fresh(entry, now: float) -> bool:
    return bool(entry and now - entry[0] < _VALUES_TTL_SECONDS)


def _subset_from_full_cache(scope: tuple[str, ...], now: float):
    if not _cache_is_fresh(_VALUES_CACHE, now):
        return None
    full_values = _VALUES_CACHE[1]
    if not all(title in full_values for title in scope):
        return None
    return {title: full_values[title] for title in scope}


def _workbook_from_values(values_by_title: dict[str, list[list]]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, values in values_by_title.items():
        worksheet = workbook.create_sheet(title)
        for row_index, row_values in enumerate(values, start=1):
            for col_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_index, column=col_index, value=value)
    if not workbook.sheetnames:
        raise RuntimeError("Google Sheets no contiene las pestañas requeridas")
    return workbook


def _remote_workbook(
    force_refresh: bool = False,
    sheet_titles: Iterable[str] | None = None,
) -> Workbook:
    global _VALUES_CACHE
    _prepare_remote_schema()
    scope = _scope_key(sheet_titles)
    now = time.monotonic()

    if not force_refresh:
        entry = _cache_entry(scope)
        if _cache_is_fresh(entry, now):
            return _workbook_from_values(entry[1])
        if scope is not None:
            from_full = _subset_from_full_cache(scope, now)
            if from_full is not None:
                _SCOPED_VALUES_CACHE[scope] = (now, from_full)
                return _workbook_from_values(from_full)

    service = _sheets_service()
    properties = _sheet_properties()
    available = {props["title"] for props in properties}
    # Nunca leer pestañas ajenas o la pestaña predeterminada `Hoja 1`.
    requested = list(scope) if scope is not None else list(REMOTE_HEADERS)
    titles = [title for title in requested if title in REMOTE_HEADERS and title in available]
    if not titles:
        raise RuntimeError("Google Sheets no contiene las pestañas requeridas para esta carga")
    try:
        response = _execute(service.spreadsheets().values().batchGet(
            spreadsheetId=_spreadsheet_id(),
            ranges=[_title_range(title) for title in titles],
            valueRenderOption="UNFORMATTED_VALUE",
            dateTimeRenderOption="FORMATTED_STRING",
        ))
    except Exception as exc:
        response_meta = getattr(exc, "resp", None)
        scoped_entry = _cache_entry(scope)
        if getattr(response_meta, "status", None) == 429:
            if scoped_entry:
                return _workbook_from_values(scoped_entry[1])
            if scope is not None:
                from_full = _subset_from_full_cache(scope, now)
                if from_full is not None:
                    return _workbook_from_values(from_full)
            if _VALUES_CACHE:
                return _workbook_from_values(_VALUES_CACHE[1])
        raise
    value_ranges = response.get("valueRanges", [])
    values_by_title: dict[str, list[list]] = {}
    for index, title in enumerate(titles):
        value_range = value_ranges[index] if index < len(value_ranges) else {}
        values_by_title[title] = value_range.get("values", [])
    entry = (now, values_by_title)
    if scope is None:
        _VALUES_CACHE = entry
        _SCOPED_VALUES_CACHE.clear()
    else:
        _SCOPED_VALUES_CACHE[scope] = entry
    return _workbook_from_values(values_by_title)


def _json_value(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _trimmed_values(ws) -> list[list]:
    values: list[list] = []
    for row in ws.iter_rows(values_only=True):
        row_values = list(row)
        while row_values and row_values[-1] in (None, ""):
            row_values.pop()
        if row_values:
            values.append([_json_value("" if value is None else value) for value in row_values])
    return values


def _snapshot(workbook: Workbook) -> dict[str, list[list]]:
    return {title: _trimmed_values(workbook[title]) for title in workbook.sheetnames}


def _max_columns(values: list[list]) -> int:
    return max((len(row) for row in values), default=0)


def _sync_one_sheet(title: str, values: list[list], previous: list[list] | None) -> dict[str, int]:
    api = _sheets_service().spreadsheets().values()
    spreadsheet_id = _spreadsheet_id()

    if previous is not None and values == previous:
        return {"rows": len(values), "cells": sum(len(row) for row in values)}

    if not values:
        _execute(api.clear(
            spreadsheetId=spreadsheet_id,
            range=_title_range(title),
            body={},
        ))
        _invalidate_values_cache()
        return {"rows": 0, "cells": 0}

    response = _execute(api.update(
        spreadsheetId=spreadsheet_id,
        range=f"{_quote_title(title)}!A1",
        valueInputOption="RAW",
        body={"values": values},
    ))

    # Primero se actualizan los datos nuevos. Si una limpieza posterior falla,
    # los datos recién registrados ya quedan persistidos en Google Sheets.
    old_rows = len(previous or [])
    old_cols = _max_columns(previous or [])
    new_rows = len(values)
    new_cols = _max_columns(values)
    max_rows = max(old_rows, new_rows, 1)
    max_cols = max(old_cols, new_cols, 1)
    if old_rows > new_rows:
        _execute(api.clear(
            spreadsheetId=spreadsheet_id,
            range=(f"{_quote_title(title)}!A{new_rows + 1}:"
                   f"{get_column_letter(max_cols)}{max_rows}"),
            body={},
        ))
    if old_cols > new_cols:
        _execute(api.clear(
            spreadsheetId=spreadsheet_id,
            range=(f"{_quote_title(title)}!{get_column_letter(new_cols + 1)}1:"
                   f"{get_column_letter(old_cols)}{max_rows}"),
            body={},
        ))

    if not response.get("updatedRange"):
        raise RuntimeError(f"Google Sheets no confirmó la actualización de '{title}'")
    _invalidate_values_cache()
    return {
        "rows": int(response.get("updatedRows", new_rows)),
        "cells": int(response.get("updatedCells", sum(len(row) for row in values))),
    }


class RemoteWorkbook:
    """Proxy de openpyxl que conoce el estado original y el alcance cargado."""

    def __init__(self, workbook: Workbook, scope: tuple[str, ...] | None = None):
        self._workbook = workbook
        self._original = _snapshot(workbook)
        self._scope = scope

    def __getitem__(self, key):
        return self._workbook[key]

    def __getattr__(self, name):
        return getattr(self._workbook, name)

    def save(self, _path: str | None = None) -> dict[str, dict[str, int]]:
        return save_workbook_to_sheets(
            self._workbook,
            original=self._original,
            cache_scope=self._scope,
        )

    def close(self) -> None:
        self._workbook.close()


def load_workbook_for_app(
    local_path: str,
    data_only: bool = False,
    read_only: bool = False,
    sheet_titles: Iterable[str] | None = None,
    force_refresh: bool = False,
):
    """Carga solo las pestañas solicitadas, o todas si no se especifican."""
    if not sheets_enabled():
        raise RuntimeError(
            "El backend remoto no está configurado. Define GOOGLE_SHEETS_ID y "
            "GOOGLE_SERVICE_ACCOUNT_JSON en Vercel."
        )
    scope = _scope_key(sheet_titles)
    return RemoteWorkbook(
        _remote_workbook(force_refresh=force_refresh, sheet_titles=scope),
        scope=scope,
    )


def save_workbook_to_sheets(
    workbook: Workbook,
    original: dict[str, list[list]] | None = None,
    cache_scope: tuple[str, ...] | None = None,
) -> dict[str, dict[str, int]]:
    """Sincroniza solo las hojas modificadas y devuelve confirmaciones de API."""
    _prepare_remote_schema()
    existing = {item["title"] for item in _sheet_properties()}
    results: dict[str, dict[str, int]] = {}
    for title in workbook.sheetnames:
        current = _trimmed_values(workbook[title])
        if title not in existing:
            _execute(_sheets_service().spreadsheets().batchUpdate(
                spreadsheetId=_spreadsheet_id(),
                body={
                    "requests": [{
                        "addSheet": {
                            "properties": {
                                "title": title,
                                "gridProperties": {
                                    "rowCount": max(1000, len(current) + 10),
                                    "columnCount": max(100, _max_columns(current)),
                                },
                            }
                        }
                    }],
                },
            ))
            existing.add(title)
        previous = original.get(title) if original else None
        results[title] = _sync_one_sheet(title, current, previous)
    snapshot = _snapshot(workbook)
    _invalidate_values_cache()
    if cache_scope is None:
        global _VALUES_CACHE
        _VALUES_CACHE = (time.monotonic(), snapshot)
    else:
        _SCOPED_VALUES_CACHE[cache_scope] = (time.monotonic(), snapshot)
    return results


def workbook_to_bytes(local_path: str) -> bytes:
    workbook = _remote_workbook()
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def export_workbook_bytes(local_paths: Iterable[str]) -> bytes:
    workbook = _remote_workbook()
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _copy_local_sheet_values(source_ws, target_ws) -> None:
    for row in source_ws.iter_rows():
        for cell in row:
            target_ws.cell(row=cell.row, column=cell.column, value=cell.value)


def _replace_workbook_contents(target: Workbook, incoming: Workbook) -> None:
    for title in list(incoming.sheetnames):
        if title in target.sheetnames:
            target.remove(target[title])
        target.create_sheet(title)
        _copy_local_sheet_values(incoming[title], target[title])
    if not target.sheetnames:
        target.create_sheet("APPTALLER")


def import_full_workbook_stream(
    stream: BinaryIO,
    local_paths: Iterable[str],
    sheet_owners: dict[str, str],
) -> list[str]:
    incoming = load_workbook(io.BytesIO(stream.read()), data_only=True)
    current = RemoteWorkbook(_remote_workbook(force_refresh=True))
    try:
        _replace_workbook_contents(current._workbook, incoming)
        result = current.save()
        return [title for title in incoming.sheetnames if title in result or title in current.sheetnames]
    finally:
        incoming.close()
        current.close()


def import_workbook_stream(
    stream: BinaryIO,
    local_path: str,
    allowed_sheets: Iterable[str] | None = None,
) -> list[str]:
    incoming = load_workbook(io.BytesIO(stream.read()), data_only=True)
    current = RemoteWorkbook(_remote_workbook(force_refresh=True))
    try:
        allowed = set(allowed_sheets) if allowed_sheets else set(incoming.sheetnames)
        for title in list(incoming.sheetnames):
            if title not in allowed:
                incoming.remove(incoming[title])
        _replace_workbook_contents(current._workbook, incoming)
        current.save()
        return list(incoming.sheetnames)
    finally:
        incoming.close()
        current.close()


def sync_seed_workbooks(local_paths: Iterable[str]) -> None:
    """Compatibilidad de API: el backend remoto no usa semillas XLSX locales."""
    _prepare_remote_schema()


def storage_status() -> dict:
    """Devuelve estado verificado sin exponer credenciales."""
    spreadsheet_id = os.getenv(SHEETS_ID_ENV, "").strip()
    try:
        info = _credentials_info()
        if not spreadsheet_id or not info:
            return {
                "backend": "google_sheets",
                "spreadsheet_configured": bool(spreadsheet_id),
                "remote_connected": False,
                "error": "Configura GOOGLE_SHEETS_ID y GOOGLE_SERVICE_ACCOUNT_JSON",
            }
        _prepare_remote_schema()
        titles = [item["title"] for item in _sheet_properties()]
        required_present = all(title in titles for title in REMOTE_HEADERS)
        return {
            "backend": "google_sheets",
            "spreadsheet_configured": True,
            "remote_connected": True,
            "spreadsheet_url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
            "sheet_titles": titles,
            "required_sheets_present": required_present,
        }
    except Exception as exc:
        return {
            "backend": "google_sheets",
            "spreadsheet_configured": bool(spreadsheet_id),
            "remote_connected": False,
            "error": f"{type(exc).__name__}: {exc}",
        }
