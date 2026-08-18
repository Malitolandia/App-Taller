#!/usr/bin/env python3
"""Backend del módulo Neveras, montado bajo /neveras en Vercel."""

from flask import Flask, jsonify, request, send_file
from werkzeug.exceptions import HTTPException
import io
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from copy import copy
from storage import load_workbook_for_app, workbook_to_bytes, google_error_status
from datetime import datetime
import os

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

@app.errorhandler(Exception)
def api_error(error):
    if isinstance(error, HTTPException):
        return error
    app.logger.exception("Error no controlado en Neveras")
    quota_status = google_error_status(error)
    if quota_status:
        return jsonify({
            "ok": False,
            "error": "Google Sheets está temporalmente limitado por cuota. Espera unos segundos y vuelve a intentar.",
            "retry_after_seconds": 10,
        }), quota_status
    return jsonify({"ok": False, "error": f"{type(error).__name__}: {error}"}), 500

# Solo se conserva el directorio de recursos de la interfaz. Los datos viven
# exclusivamente en Google Sheets y nunca se buscan en un Excel local.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = "google-sheets://Neveras"

# ─────────────────────────────────────────────────────────────
# COLUMNAS DE LA HOJA "Ventas"
#
#  A(1)  #             ← FORMULA  — NO TOCAR
#  B(2)  Fecha         ← escribir
#  C(3)  Hora          ← escribir
#  D(4)  Cliente       ← escribir
#  E(5)  Producto      ← escribir
#  F(6)  Cantidad      ← escribir
#  G(7)  Precio Unit.  ← FORMULA  — NO TOCAR
#  H(8)  Total         ← FORMULA  — NO TOCAR
#  I(9)  Método Pago   ← escribir
#  J(10) Pagó          ← escribir
#  K(11) Estado Pago   ← escribir
#  L(12) Ganancia      ← FORMULA  — NO TOCAR
#
# La fila vacía se detecta mirando la columna B (no la A).
# ─────────────────────────────────────────────────────────────
COLS_DATOS    = [2, 3, 4, 5, 6, 9, 10, 11]   # B C D E F I J K
COLS_FORMULA  = [1, 7, 8, 12]                  # A G H L  — intocables


def _at(row, index, default=None):
    """Acceso tolerante a filas remotas con celdas vacías finales."""
    return row[index] if index < len(row) else default


def _as_float(value, default=0.0):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return default


def _as_int(value, default=0):
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return default


# ── UTILIDAD: copiar formato de una celda a otra ─────────────

def _copiar_formato(origen, destino):
    """Copia font, fill, border, alignment y number_format de origen a destino."""
    if origen.font:
        destino.font       = copy(origen.font)
    if origen.fill:
        destino.fill       = copy(origen.fill)
    if origen.border:
        destino.border     = copy(origen.border)
    if origen.alignment:
        destino.alignment  = copy(origen.alignment)
    destino.number_format  = origen.number_format


# ── LECTURA ──────────────────────────────────────────────────

def _precio_de_inventario(producto):
    """Devuelve (precio, ganUnit) desde la hoja Inventario."""
    try:
        wb = load_workbook_for_app(EXCEL_PATH, data_only=True, read_only=True)
        ws = wb['Inventario']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _at(row, 0) and str(_at(row, 0)) == producto:
                precio = _as_float(_at(row, 2))
                costo  = _as_float(_at(row, 1))
                wb.close()
                return precio, precio - costo
        wb.close()
    except Exception:
        pass
    return 0, 0


def leer_ventas(prods_base=None):
    """
    Lee la hoja Ventas.
    Para filas donde Total=0 (fórmula aún sin calcular) rellena
    precio y total desde el inventario base para que el dashboard
    muestre valores correctos inmediatamente.
    """
    wb     = load_workbook_for_app(EXCEL_PATH, data_only=True, read_only=True)
    ws     = wb['Ventas']
    ventas = []

    # Índice rápido precio → para rellenar filas recién escritas
    precio_idx = {}
    if prods_base:
        for p in prods_base:
            precio_idx[p['producto']] = (p['precio'], p['precio'] - p['costo'])

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Detectar fila vacía por columna B (índice 1)
        if not _at(row, 1):
            continue
        num      = _at(row, 0)
        cantidad = _as_int(_at(row, 5))
        producto = str(_at(row, 4) or '')
        precio   = _as_float(_at(row, 6))
        total    = _as_float(_at(row, 7))
        ganancia = _as_float(_at(row, 11))

        # Si las fórmulas aún no tienen valor cacheado, calcular desde inventario
        if precio == 0 and producto in precio_idx:
            precio, ganUnit = precio_idx[producto]
            total    = cantidad * precio
            ganancia = round(cantidad * ganUnit, 2)

        ventas.append({
            'num':      int(num) if num else len(ventas) + 1,
                        'fecha':     str(_at(row, 1) or '')[:10],

                        'hora':      str(_at(row, 2) or ''),
            'cliente':   str(_at(row, 3) or '').upper(),

            'producto': producto,
            'cantidad': cantidad,
            'precio':   precio,
            'total':    total,
                        'metodo':    str(_at(row, 8) or ''),
            'pago':      'SI' if str(_at(row, 9) or '').upper() == 'SI' else 'NO',
            'estado':    str(_at(row, 10) or ''),

            'ganancia': ganancia,
        })

    wb.close()
    return ventas


def leer_inventario_base():
    wb       = load_workbook_for_app(EXCEL_PATH, data_only=True, read_only=True)
    ws       = wb['Inventario']
    prods    = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not _at(row, 0):
            continue
        prods.append({
            'producto': str(_at(row, 0)),
            'costo':    _as_float(_at(row, 1)),
            'precio':   _as_float(_at(row, 2)),
            'stockIni': _as_float(_at(row, 4)),
            'stockMin': _as_float(_at(row, 5)),
        })
    wb.close()
    return prods


def _primera_fila_vacia_inventario(ws):
    """Devuelve la primera fila libre de Inventario según la columna A."""
    max_row = max(int(ws.max_row or 1), 1)
    for row_number in range(2, max_row + 1):
        value = ws.cell(row=row_number, column=1).value
        if value is None or str(value).strip() == '':
            return row_number
    return max_row + 1


def _numero_no_negativo(value, campo, entero=False):
    """Valida importes y cantidades recibidos desde el formulario web."""
    if value is None or str(value).strip() == '':
        raise ValueError(f'El campo {campo} es obligatorio')
    try:
        numero = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'El campo {campo} debe ser numérico') from exc
    if numero != numero or numero in (float('inf'), float('-inf')) or numero < 0:
        raise ValueError(f'El campo {campo} debe ser un número no negativo')
    if entero and not numero.is_integer():
        raise ValueError(f'El campo {campo} debe ser un número entero')
    return int(numero) if entero else round(numero, 2)


def agregar_producto(producto, costo, precio, stock_inicial, stock_minimo):
    """Agrega un producto nuevo a Inventario y confirma la sincronización remota."""
    wb = None
    try:
        nombre = str(producto or '').strip()
        if not nombre:
            raise ValueError('El nombre del producto es obligatorio')

        costo = _numero_no_negativo(costo, 'Costo')
        precio = _numero_no_negativo(precio, 'Precio de venta')
        stock_inicial = _numero_no_negativo(stock_inicial, 'Stock inicial', entero=True)
        stock_minimo = _numero_no_negativo(stock_minimo, 'Stock mínimo', entero=True)

        wb = load_workbook_for_app(EXCEL_PATH)
        ws = wb['Inventario']
        nombres_existentes = {
            str(ws.cell(row=row_number, column=1).value).strip().casefold()
            for row_number in range(2, max(int(ws.max_row or 1), 1) + 1)
            if ws.cell(row=row_number, column=1).value not in (None, '')
        }
        if nombre.casefold() in nombres_existentes:
            return False, 'Ya existe un producto con ese nombre', None

        fila = _primera_fila_vacia_inventario(ws)
        ganancia_unitaria = round(precio - costo, 2)
        stock_actual = stock_inicial
        estado = '🚫 AGOTADO' if stock_actual <= 0 else ('⚠ SURTIR' if stock_actual <= stock_minimo else '✅ OK')
        valores = [
            nombre,
            costo,
            precio,
            ganancia_unitaria,
            stock_inicial,
            stock_minimo,
            0,
            stock_actual,
            estado,
            round(costo * stock_inicial, 2),
            round(ganancia_unitaria * stock_inicial, 2),
        ]
        for column, value in enumerate(valores, start=1):
            ws.cell(row=fila, column=column, value=value)

        resultado = wb.save(EXCEL_PATH)
        confirmado = bool(resultado and resultado.get('Inventario'))
        if not confirmado:
            raise RuntimeError('Google Sheets no confirmó la actualización de Inventario')

        return True, None, {
            'producto': nombre,
            'costo': costo,
            'precio': precio,
            'ganUnit': ganancia_unitaria,
            'stockIni': stock_inicial,
            'stockMin': stock_minimo,
            'vendidos': 0,
            'stockAct': stock_actual,
            'estado': estado,
        }
    finally:
        if wb is not None:
            wb.close()


def _fila_inventario(ws, producto):
    """Busca un producto por nombre sin distinguir mayúsculas/minúsculas."""
    clave = str(producto or '').strip().casefold()
    if not clave:
        return None
    for fila in range(2, max(int(ws.max_row or 1), 1) + 1):
        valor = ws.cell(row=fila, column=1).value
        if valor is not None and str(valor).strip().casefold() == clave:
            return fila
    return None


def _unidades_vendidas_en_libro(ws_ventas, producto):
    """Cuenta unidades vendidas de un producto desde la hoja Ventas."""
    clave = str(producto or '').strip().casefold()
    vendidas = 0
    for fila in range(2, max(int(ws_ventas.max_row or 1), 1) + 1):
        if ws_ventas.cell(row=fila, column=2).value in (None, ''):
            continue
        nombre = str(ws_ventas.cell(row=fila, column=5).value or '').strip().casefold()
        if nombre == clave:
            vendidas += _as_int(ws_ventas.cell(row=fila, column=6).value)
    return vendidas


def _congelar_ventas_historicas(ws_ventas, producto, precio, costo):
    """Conserva precio, total y ganancia de ventas anteriores a la edición."""
    clave = str(producto or '').strip().casefold()
    gan_unit = round(precio - costo, 2)
    for fila in range(2, max(int(ws_ventas.max_row or 1), 1) + 1):
        if ws_ventas.cell(row=fila, column=2).value in (None, ''):
            continue
        nombre = str(ws_ventas.cell(row=fila, column=5).value or '').strip().casefold()
        if nombre != clave:
            continue
        cantidad = _as_int(ws_ventas.cell(row=fila, column=6).value)
        ws_ventas.cell(row=fila, column=7).value = precio
        ws_ventas.cell(row=fila, column=8).value = round(cantidad * precio, 2)
        ws_ventas.cell(row=fila, column=12).value = round(cantidad * gan_unit, 2)


def _escribir_resumen_inventario(ws, fila, producto, costo, precio, stock_inicial, stock_minimo, vendidos):
    """Escribe las columnas de Inventario y devuelve el resumen calculado."""
    gan_unit = round(precio - costo, 2)
    stock_actual = stock_inicial - vendidos
    if stock_actual <= 0:
        estado = '🚫 AGOTADO'
    elif stock_actual <= stock_minimo:
        estado = '⚠ SURTIR'
    else:
        estado = '✅ OK'

    valores = {
        1: producto,
        2: costo,
        3: precio,
        4: gan_unit,
        5: stock_inicial,
        6: stock_minimo,
        7: vendidos,
        8: stock_actual,
        9: estado,
        10: round(costo * stock_inicial, 2),
        11: round(gan_unit * stock_inicial, 2),
    }
    for columna, valor in valores.items():
        ws.cell(row=fila, column=columna).value = valor

    return {
        'producto': producto,
        'costo': costo,
        'precio': precio,
        'ganUnit': gan_unit,
        'stockIni': stock_inicial,
        'stockMin': stock_minimo,
        'vendidos': vendidos,
        'stockAct': stock_actual,
        'estado': estado,
    }


def _validar_datos_producto(producto, costo, precio, stock_inicial, stock_minimo):
    nombre = str(producto or '').strip()
    if not nombre:
        raise ValueError('El nombre del producto es obligatorio')
    return {
        'producto': nombre,
        'costo': _numero_no_negativo(costo, 'Costo'),
        'precio': _numero_no_negativo(precio, 'Precio de venta'),
        'stockInicial': _numero_no_negativo(stock_inicial, 'Stock inicial', entero=True),
        'stockMin': _numero_no_negativo(stock_minimo, 'Stock mínimo', entero=True),
    }


def editar_producto(producto_original, producto, costo, precio, stock_inicial, stock_minimo):
    """Edita los datos de Inventario sin romper ventas históricas."""
    datos = _validar_datos_producto(producto, costo, precio, stock_inicial, stock_minimo)
    original = str(producto_original or '').strip()
    if not original:
        raise ValueError('El producto original es obligatorio')

    wb = None
    try:
        wb = load_workbook_for_app(EXCEL_PATH)
        ws = wb['Inventario']
        ws_ventas = wb['Ventas']
        fila = _fila_inventario(ws, original)
        if fila is None:
            return False, f"No se encontró el producto '{original}'", None

        fila_otro = _fila_inventario(ws, datos['producto'])
        if fila_otro is not None and fila_otro != fila:
            return False, 'Ya existe otro producto con ese nombre', None

        vendidos = _unidades_vendidas_en_libro(ws_ventas, original)
        if datos['producto'].casefold() != original.casefold() and vendidos > 0:
            return False, 'No se puede cambiar el nombre de un producto con ventas históricas', None

        costo_anterior = _as_float(ws.cell(row=fila, column=2).value)
        precio_anterior = _as_float(ws.cell(row=fila, column=3).value)
        if vendidos > 0 and (
            abs(datos['costo'] - costo_anterior) > 0.005
            or abs(datos['precio'] - precio_anterior) > 0.005
        ):
            _congelar_ventas_historicas(ws_ventas, original, precio_anterior, costo_anterior)

        resumen = _escribir_resumen_inventario(
            ws, fila, datos['producto'], datos['costo'], datos['precio'],
            datos['stockInicial'], datos['stockMin'], vendidos,
        )
        resultado = wb.save(EXCEL_PATH)
        if not resultado or not resultado.get('Inventario'):
            raise RuntimeError('Google Sheets no confirmó la edición de Inventario')
        return True, f"✅ Producto '{datos['producto']}' actualizado en Google Sheets", resumen
    finally:
        if wb is not None:
            wb.close()


def ajustar_existencias(producto, stock_actual):
    """Fija la existencia actual conservando el histórico de unidades vendidas."""
    nombre = str(producto or '').strip()
    if not nombre:
        raise ValueError('El nombre del producto es obligatorio')
    nuevo_stock = _numero_no_negativo(stock_actual, 'Nueva existencia', entero=True)

    wb = None
    try:
        wb = load_workbook_for_app(EXCEL_PATH)
        ws = wb['Inventario']
        ws_ventas = wb['Ventas']
        fila = _fila_inventario(ws, nombre)
        if fila is None:
            return False, f"No se encontró el producto '{nombre}'", None

        vendidos = _unidades_vendidas_en_libro(ws_ventas, nombre)
        costo = _as_float(ws.cell(row=fila, column=2).value)
        precio = _as_float(ws.cell(row=fila, column=3).value)
        stock_minimo = _as_int(ws.cell(row=fila, column=6).value)
        # Stock Inicial = existencia actual deseada + histórico vendido.
        stock_inicial = nuevo_stock + vendidos
        resumen = _escribir_resumen_inventario(
            ws, fila, str(ws.cell(row=fila, column=1).value).strip(),
            costo, precio, stock_inicial, stock_minimo, vendidos,
        )
        resultado = wb.save(EXCEL_PATH)
        if not resultado or not resultado.get('Inventario'):
            raise RuntimeError('Google Sheets no confirmó el ajuste de existencias')
        return True, f"✅ Existencias de '{nombre}' actualizadas a {nuevo_stock}", resumen
    finally:
        if wb is not None:
            wb.close()


def eliminar_producto(producto):
    """Elimina un producto solo si no tiene ventas históricas asociadas."""
    nombre = str(producto or '').strip()
    if not nombre:
        raise ValueError('El nombre del producto es obligatorio')

    wb = None
    try:
        wb = load_workbook_for_app(EXCEL_PATH)
        ws = wb['Inventario']
        ws_ventas = wb['Ventas']
        fila = _fila_inventario(ws, nombre)
        if fila is None:
            return False, f"No se encontró el producto '{nombre}'", None

        vendidos = _unidades_vendidas_en_libro(ws_ventas, nombre)
        if vendidos > 0:
            return False, (
                f"No se puede eliminar '{nombre}' porque tiene {vendidos} "
                'unidades vendidas. Puedes editarlo o ajustar sus existencias.'
            ), None

        ws.delete_rows(fila, 1)
        resultado = wb.save(EXCEL_PATH)
        if not resultado or not resultado.get('Inventario'):
            raise RuntimeError('Google Sheets no confirmó la eliminación de Inventario')
        return True, f"✅ Producto '{nombre}' eliminado de Google Sheets", {'producto': nombre}
    finally:
        if wb is not None:
            wb.close()


def calcular_inventario_y_clientes(ventas, prods_base):
    inventario = []
    for p in prods_base:
        vendidos  = sum(v['cantidad'] for v in ventas if v['producto'] == p['producto'])
        stock_act = p['stockIni'] - vendidos
        ganUnit   = p['precio'] - p['costo']
        if stock_act <= 0:
            estado = '🚫 AGOTADO'
        elif stock_act <= p['stockMin']:
            estado = '⚠ SURTIR'
        else:
            estado = '✅ OK'
        inventario.append({
            'producto': p['producto'],
            'costo':    p['costo'],
            'precio':   p['precio'],
            'ganUnit':  ganUnit,
            'stockIni': p['stockIni'],
            'stockMin': p['stockMin'],
            'vendidos': vendidos,
            'stockAct': stock_act,
            'estado':   estado,
        })

    cli_dict = {}
    for v in ventas:
        cli = v['cliente']
        if not cli:
            continue
        if cli not in cli_dict:
            cli_dict[cli] = {'comprado': 0, 'pagado': 0, 'compras': 0}
        cli_dict[cli]['comprado'] += v['total']
        if v['pago'] == 'SI':
            cli_dict[cli]['pagado'] += v['total']
        cli_dict[cli]['compras'] += 1

    clientes = [{
        'cliente':  cli,
        'comprado': v['comprado'],
        'pagado':   v['pagado'],
        'deuda':    v['comprado'] - v['pagado'],
        'compras':  v['compras'],
    } for cli, v in cli_dict.items()]

    return inventario, clientes


def _datos_actualizados_neveras():
    prods = leer_inventario_base()
    ventas = leer_ventas(prods)
    inventario, clientes = calcular_inventario_y_clientes(ventas, prods)
    return ventas, inventario, clientes


# ── ESCRITURA ─────────────────────────────────────────────────

def _primera_fila_vacia(ws):
    """
    Devuelve el número de la primera fila donde la columna B esté vacía.
    Se ignora la columna A porque tiene fórmula y puede tener valor
    incluso si la fila no fue rellenada manualmente.
    """
    fila = 2
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        celda = row[0]
        if celda.value is None or str(celda.value).strip() == '':
            return celda.row
        fila = celda.row + 1
    return fila


def agregar_fila_venta(cliente, producto, cantidad, metodo, pago):
    """
    Agrega una fila en la hoja 'Ventas'.

    REGLAS:
    - Solo escribe en cols B C D E F I J K (índices 2 3 4 5 6 9 10 11).
    - Nunca toca A (formula #), G (Precio), H (Total), L (Ganancia).
    - La fila vacía se localiza mirando la columna B.
    - Copia el formato visual de la fila anterior a las celdas nuevas.
    """
    try:
        # Cargar SIN data_only para preservar las fórmulas existentes
        wb = load_workbook_for_app(EXCEL_PATH)
        ws = wb['Ventas']

        fila      = _primera_fila_vacia(ws)
        fila_ref  = fila - 1   # fila de referencia para copiar formato

        ahora  = datetime.now()
        estado = '✅ PAGADO' if pago.upper() == 'SI' else '🔴 PENDIENTE'

        # Valores a escribir indexados por número de columna
        valores = {
            2:  ahora.strftime('%Y-%m-%d'),   # Fecha
            3:  ahora.strftime('%H:%M'),       # Hora
            4:  cliente.upper(),              # Cliente
            5:  producto,                     # Producto
            6:  cantidad,                     # Cantidad
            9:  metodo,                       # Método Pago
            10: pago.upper(),                 # Pagó
            11: estado,                       # Estado Pago
        }

        for col, valor in valores.items():
            celda_nueva = ws.cell(row=fila,     column=col)
            celda_ref   = ws.cell(row=fila_ref, column=col)

            # Escribir valor
            celda_nueva.value = valor

            # Copiar formato de la fila anterior (si existe)
            if fila_ref >= 2:
                _copiar_formato(celda_ref, celda_nueva)

        wb.save(EXCEL_PATH)
        wb.close()
        return True, None

    except PermissionError:
        return False, ('El archivo Excel está abierto. '
                       'Ciérralo en Excel y vuelve a intentarlo.')
    except Exception as e:
        return False, str(e)


def agregar_filas_venta(cliente, items, metodo, pago):
    """
    Agrega varias filas en la hoja 'Ventas' (una por cada producto de la
    compra), todas con el mismo cliente, método de pago, estado de pago,
    fecha y hora — para representar una sola compra con varios productos.

    items: lista de tuplas (producto, cantidad).
    Misma lógica/columnas que agregar_fila_venta, pero abre y guarda el
    Excel una sola vez para todas las filas.
    """
    try:
        wb = load_workbook_for_app(EXCEL_PATH)
        ws = wb['Ventas']

        ahora  = datetime.now()
        estado = '✅ PAGADO' if pago.upper() == 'SI' else '🔴 PENDIENTE'

        for producto, cantidad in items:
            fila     = _primera_fila_vacia(ws)
            fila_ref = fila - 1

            valores = {
                2:  ahora.strftime('%Y-%m-%d'),   # Fecha
                3:  ahora.strftime('%H:%M'),       # Hora
                4:  cliente.upper(),              # Cliente
                5:  producto,                     # Producto
                6:  cantidad,                     # Cantidad
                9:  metodo,                       # Método Pago
                10: pago.upper(),                 # Pagó
                11: estado,                       # Estado Pago
            }

            for col, valor in valores.items():
                celda_nueva = ws.cell(row=fila,     column=col)
                celda_ref   = ws.cell(row=fila_ref, column=col)
                celda_nueva.value = valor
                if fila_ref >= 2:
                    _copiar_formato(celda_ref, celda_nueva)

        wb.save(EXCEL_PATH)
        wb.close()
        return True, None

    except PermissionError:
        return False, ('El archivo Excel está abierto. '
                       'Ciérralo en Excel y vuelve a intentarlo.')
    except Exception as e:
        return False, str(e)


def marcar_venta_pagada(num):
    """Marca la venta como pagada. Solo toca cols J y K."""
    try:
        wb = load_workbook_for_app(EXCEL_PATH)
        ws = wb['Ventas']

        # ── CORRECCIÓN ────────────────────────────────────────────
        # La col A tiene una fórmula (p. ej. =ROW()-1).
        # Al abrir sin data_only=True, row[0].value devuelve el STRING
        # de la fórmula, nunca el número → la comparación siempre fallaba.
        #
        # Solución: calcular la fila directamente.
        # Fila 1 = encabezado → venta #N siempre está en fila N + 1.
        # ─────────────────────────────────────────────────────────
        fila = int(num) + 1

        # Verificar que la fila tiene datos (col B = Fecha)
        if ws.cell(row=fila, column=2).value is None:
            wb.close()
            return False, f'Venta #{num} no encontrada'

        ws.cell(row=fila, column=10).value = 'SI'
        ws.cell(row=fila, column=11).value = '✅ PAGADO'
        wb.save(EXCEL_PATH)
        wb.close()
        return True, None

    except PermissionError:
        return False, 'El archivo Excel está abierto. Ciérralo e intenta de nuevo.'
    except Exception as e:
        return False, str(e)


def cobrar_deuda_cliente(cliente, monto):
    """Cobra ventas pendientes completas de un cliente, total o parcialmente.

    El cobro parcial se aplica en orden cronológico y nunca divide una venta:
    se marcan tantas ventas completas como permita el monto recibido. Si el
    importe es menor que la primera venta pendiente, se devuelve un error para
    evitar registrar un cobro ambiguo.
    """
    nombre = str(cliente or '').strip().upper()
    if not nombre:
        return False, 'El cliente es obligatorio', None

    try:
        monto = float(monto)
    except (TypeError, ValueError):
        return False, 'El monto debe ser numérico', None
    if monto != monto or monto in (float('inf'), float('-inf')) or monto <= 0:
        return False, 'El monto debe ser mayor que cero', None
    monto = round(monto, 2)

    wb = None
    try:
        # Una sola carga lógica del libro y un solo save al finalizar.
        wb = load_workbook_for_app(EXCEL_PATH)
        ws_ventas = wb['Ventas']
        ws_inv = wb['Inventario']

        precios = {}
        for fila in ws_inv.iter_rows(min_row=2, values_only=True):
            producto = _at(fila, 0)
            if producto:
                precios[str(producto).strip()] = _as_float(_at(fila, 2))

        pendientes = []
        deuda_total = 0.0
        for fila_num in range(2, max(int(ws_ventas.max_row or 1), 1) + 1):
            fecha = ws_ventas.cell(row=fila_num, column=2).value
            if fecha is None or str(fecha).strip() == '':
                continue

            cli = str(ws_ventas.cell(row=fila_num, column=4).value or '').strip().upper()
            pago = str(ws_ventas.cell(row=fila_num, column=10).value or '').strip().upper()
            if cli != nombre or pago == 'SI':
                continue

            producto = str(ws_ventas.cell(row=fila_num, column=5).value or '').strip()
            cantidad = _as_int(ws_ventas.cell(row=fila_num, column=6).value)
            precio = _as_float(ws_ventas.cell(row=fila_num, column=7).value)
            total = _as_float(ws_ventas.cell(row=fila_num, column=8).value)
            # Las fórmulas pueden no traer valor cacheado inmediatamente;
            # calcular el total con el inventario como ya hace leer_ventas().
            if total <= 0 and producto in precios:
                precio = precios[producto]
                total = round(cantidad * precio, 2)
            if total <= 0:
                continue

            numero = ws_ventas.cell(row=fila_num, column=1).value
            try:
                numero = int(float(numero))
            except (TypeError, ValueError):
                numero = fila_num - 1

            pendientes.append({'fila': fila_num, 'num': numero, 'total': total})
            deuda_total += total

        deuda_total = round(deuda_total, 2)
        if not pendientes:
            return False, f'El cliente {nombre} no tiene deudas pendientes', None

        # Cobro total o selección de ventas completas hasta agotar el monto.
        seleccionadas = []
        acumulado = 0.0
        for venta in pendientes:
            if monto >= deuda_total - 0.005:
                seleccionadas.append(venta)
                continue
            if acumulado + venta['total'] <= monto + 0.005:
                seleccionadas.append(venta)
                acumulado += venta['total']
            else:
                break

        if not seleccionadas:
            primera = pendientes[0]['total']
            return False, (
                f'El monto es menor que la primera venta pendiente '
                f'({primera:.2f}). Para cobro parcial, ingresa al menos ese valor.'
            ), None

        for venta in seleccionadas:
            ws_ventas.cell(row=venta['fila'], column=10).value = 'SI'
            ws_ventas.cell(row=venta['fila'], column=11).value = '✅ PAGADO'

        monto_aplicado = round(sum(v['total'] for v in seleccionadas), 2)
        deuda_restante = round(max(0.0, deuda_total - monto_aplicado), 2)
        resultado = wb.save(EXCEL_PATH)
        if not resultado or not resultado.get('Ventas'):
            raise RuntimeError('Google Sheets no confirmó la actualización de Ventas')

        detalle = {
            'cliente': nombre,
            'montoSolicitado': monto,
            'montoAplicado': monto_aplicado,
            'deudaAnterior': deuda_total,
            'deudaRestante': deuda_restante,
            'ventasCobradas': [v['num'] for v in seleccionadas],
        }
        modalidad = 'total' if deuda_restante <= 0.005 else 'parcial'
        mensaje = (
            f'✅ Cobro {modalidad} aplicado a {nombre}: '
            f'${monto_aplicado:,.2f}'
        )
        return True, mensaje, detalle

    except PermissionError:
        return False, 'El archivo Excel está abierto. Ciérralo e intenta de nuevo.', None
    except Exception:
        # Dejar que el manejador global preserve, entre otros, el 503 de cuota.
        raise
    finally:
        if wb is not None:
            wb.close()


# ── HELPER: archivos estáticos en ambas carpetas ─────────────

def _static(nombre):
    return os.path.join(BASE_DIR, nombre)


# ── RUTAS ─────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_file(_static('dashboard_neveras.html'))

@app.route('/styles.css')
def styles():
    return send_file(_static('styles.css'))

@app.route('/app.js')
def appjs():
    return send_file(_static('app.js'))


@app.route('/api/datos')
def get_datos():
    try:
        prods  = leer_inventario_base()
        ventas = leer_ventas(prods)
        inv, cli = calcular_inventario_y_clientes(ventas, prods)
        return jsonify({'ventas': ventas, 'inventario': inv, 'clientes': cli})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/nueva-venta', methods=['POST'])
def nueva_venta():
    data    = request.json
    cliente = data.get('cliente', '').strip()
    metodo  = data.get('metodo', 'Efectivo')
    pago    = data.get('pago', 'NO').upper()

    # Nuevo formato: varios productos en una sola compra.
    # Se mantiene compatibilidad con el formato anterior (un solo producto).
    items_in = data.get('items')
    if items_in:
        items = [(it.get('producto', ''), int(it.get('cantidad', 0) or 0)) for it in items_in]
    else:
        items = [(data.get('producto', ''), int(data.get('cantidad', 1) or 0))]
    items = [(p, c) for p, c in items if p and c >= 1]

    if not cliente or not items:
        return jsonify({'error': 'Datos inválidos'}), 400

    ok, err = agregar_filas_venta(cliente, items, metodo, pago)
    if not ok:
        return jsonify({'error': err}), 500

    prods  = leer_inventario_base()
    ventas = leer_ventas(prods)
    inv, cli = calcular_inventario_y_clientes(ventas, prods)
    resumen = ', '.join(f'{p} x{c}' for p, c in items)
    return jsonify({
        'ok':         True,
        'mensaje':    f'✅ Venta guardada: {cliente} — {resumen}',
        'ventas':     ventas,
        'inventario': inv,
        'clientes':   cli,
    })


@app.route('/api/marcar-pagado', methods=['POST'])
def marcar_pagado():
    data = request.json
    num  = data.get('num')
    if not num:
        return jsonify({'error': 'Número de venta requerido'}), 400

    ok, err = marcar_venta_pagada(int(num))
    if not ok:
        return jsonify({'error': err}), 500

    prods  = leer_inventario_base()
    ventas = leer_ventas(prods)
    inv, cli = calcular_inventario_y_clientes(ventas, prods)
    return jsonify({
        'ok':         True,
        'mensaje':    f'✅ Venta #{num} marcada como pagada',
        'ventas':     ventas,
        'inventario': inv,
        'clientes':   cli,
    })


@app.route('/api/cobrar-cliente', methods=['POST'])
def cobrar_cliente():
    data = request.get_json(silent=True) or {}
    cliente = data.get('cliente', '')
    monto = data.get('monto')
    ok, mensaje, detalle = cobrar_deuda_cliente(cliente, monto)
    if not ok:
        return jsonify({'ok': False, 'error': mensaje}), 400

    prods  = leer_inventario_base()
    ventas = leer_ventas(prods)
    inv, cli = calcular_inventario_y_clientes(ventas, prods)
    return jsonify({
        'ok':         True,
        'mensaje':    mensaje,
        'detalle':    detalle,
        'ventas':     ventas,
        'inventario': inv,
        'clientes':   cli,
    })


@app.route('/api/productos')
def get_productos():
    try:
        prods  = leer_inventario_base()
        ventas = leer_ventas(prods)
        inv, _ = calcular_inventario_y_clientes(ventas, prods)
        return jsonify([{
            'nombre':  p['producto'],
            'precio':  p['precio'],
            'ganUnit': p['ganUnit'],
            'stock':   p['stockAct'],
        } for p in inv])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/nuevo-producto', methods=['POST'])
def nuevo_producto():
    data = request.get_json(silent=True) or {}
    try:
        ok, error, producto = agregar_producto(
            data.get('producto'),
            data.get('costo'),
            data.get('precio'),
            data.get('stockInicial'),
            data.get('stockMin'),
        )
        if not ok:
            return jsonify({'ok': False, 'error': error}), 409
        return jsonify({
            'ok': True,
            'mensaje': f"✅ Producto '{producto['producto']}' guardado en Google Sheets",
            'producto': producto,
        })
    except Exception as exc:
        quota_status = google_error_status(exc)
        if quota_status:
            return jsonify({
                'ok': False,
                'error': 'Google Sheets está temporalmente limitado por cuota. Espera unos segundos y vuelve a intentar.',
                'retry_after_seconds': 10,
            }), quota_status
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/editar-producto', methods=['POST'])
def editar_producto_api():
    data = request.get_json(silent=True) or {}
    try:
        ok, mensaje, producto = editar_producto(
            data.get('productoOriginal'),
            data.get('producto'),
            data.get('costo'),
            data.get('precio'),
            data.get('stockInicial'),
            data.get('stockMin'),
        )
        if not ok:
            return jsonify({'ok': False, 'error': mensaje}), 409
        ventas, inventario, clientes = _datos_actualizados_neveras()
        return jsonify({
            'ok': True,
            'mensaje': mensaje,
            'producto': producto,
            'ventas': ventas,
            'inventario': inventario,
            'clientes': clientes,
        })
    except Exception as exc:
        quota_status = google_error_status(exc)
        if quota_status:
            return jsonify({
                'ok': False,
                'error': 'Google Sheets está temporalmente limitado por cuota. Espera unos segundos y vuelve a intentar.',
                'retry_after_seconds': 10,
            }), quota_status
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/ajustar-existencias', methods=['POST'])
def ajustar_existencias_api():
    data = request.get_json(silent=True) or {}
    try:
        ok, mensaje, producto = ajustar_existencias(
            data.get('producto'),
            data.get('stockActual'),
        )
        if not ok:
            return jsonify({'ok': False, 'error': mensaje}), 409
        ventas, inventario, clientes = _datos_actualizados_neveras()
        return jsonify({
            'ok': True,
            'mensaje': mensaje,
            'producto': producto,
            'ventas': ventas,
            'inventario': inventario,
            'clientes': clientes,
        })
    except Exception as exc:
        quota_status = google_error_status(exc)
        if quota_status:
            return jsonify({
                'ok': False,
                'error': 'Google Sheets está temporalmente limitado por cuota. Espera unos segundos y vuelve a intentar.',
                'retry_after_seconds': 10,
            }), quota_status
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/eliminar-producto', methods=['POST'])
def eliminar_producto_api():
    data = request.get_json(silent=True) or {}
    try:
        ok, mensaje, producto = eliminar_producto(data.get('producto'))
        if not ok:
            return jsonify({'ok': False, 'error': mensaje}), 409
        ventas, inventario, clientes = _datos_actualizados_neveras()
        return jsonify({
            'ok': True,
            'mensaje': mensaje,
            'producto': producto,
            'ventas': ventas,
            'inventario': inventario,
            'clientes': clientes,
        })
    except Exception as exc:
        quota_status = google_error_status(exc)
        if quota_status:
            return jsonify({
                'ok': False,
                'error': 'Google Sheets está temporalmente limitado por cuota. Espera unos segundos y vuelve a intentar.',
                'retry_after_seconds': 10,
            }), quota_status
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/lista-clientes')
def lista_clientes():
    """
    Lee TODA la columna A de la hoja 'Clientes' desde la fila 2.
    No usa read_only para evitar que openpyxl corte el max_row antes de tiempo.
    """
    try:
        wb = load_workbook_for_app(EXCEL_PATH, data_only=True)
        ws = wb['Clientes']
        nombres = []
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            val = row[0].value
            if val is not None and str(val).strip() != '':
                nombres.append(str(val).strip().upper())
        wb.close()
        return jsonify(nombres)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/excel')
def download_excel():
    content = workbook_to_bytes(EXCEL_PATH)
    return send_file(io.BytesIO(content),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='INVENTARIO_NEVERAS.xlsx')


# Vercel carga este módulo desde el WSGI unificado de ../app.py.
